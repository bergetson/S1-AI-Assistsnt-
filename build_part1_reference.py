"""
Part 1: Create workbook skeleton + all [ADMIN] reference sheets with demo data.
Run: python3 build_part1_reference.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME
import openpyxl

# ── Color Palette ──────────────────────────────────────────────────────────────
ARMY_GREEN      = "1B4F2A"
ARMY_TAN        = "C8A96E"
ARMY_BLACK      = "1A1A1A"
HEADER_BLUE     = "1F3864"
LIGHT_BLUE      = "D9E1F2"
LIGHT_GREEN     = "E2EFDA"
LIGHT_YELLOW    = "FFFF99"
LIGHT_ORANGE    = "FCE4D6"
LIGHT_RED       = "FFE0E0"
WHITE           = "FFFFFF"
LIGHT_GRAY      = "F2F2F2"
MED_GRAY        = "D9D9D9"
DEMO_ORANGE     = "FF6600"   # used to flag demo data

# ── Style Helpers ──────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color=ARMY_BLACK):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def style_header_row(ws, row, cols, bg=HEADER_BLUE, fg=WHITE, size=10):
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.font = header_font(size=size, color=fg)
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = thin_border()

def style_data_row(ws, row, cols, bg=WHITE, fg=ARMY_BLACK):
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.font = body_font(color=fg)
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()

def section_title(ws, row, col, text, bg=ARMY_GREEN, fg=WHITE, span=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = header_font(size=12, color=fg)
    cell.fill = fill(bg)
    cell.alignment = center()
    cell.border = thin_border()

def write_header(ws, row, headers, bg=HEADER_BLUE):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = header_font(size=9)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def write_row(ws, row, values, bg=WHITE, fg=ARMY_BLACK, bold=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name="Calibri", size=9, bold=bold, color=fg)
        c.fill = fill(bg)
        c.alignment = left()
        c.border = thin_border()

# ── Build Workbook ─────────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ── Sheet order ───────────────────────────────────────────────────────────────
sheet_names = [
    "HOME",
    "MY PROFILE",
    "POSITION MATCHES",
    "COMMUTE ANALYSIS",
    "CAREER TIMELINE",
    "COUNSELING SHEET",
    "ADMIN - POSITIONS",
    "ADMIN - CITY DISTANCES",
    "ADMIN - PME AND SCHOOLS",
    "ADMIN - MOS TRANSITIONS",
    "ADMIN - PROMOTION TIMELINE",
    "ADMIN - CONFIG",
    "CALC ENGINE",
]

sheets = {}
for name in sheet_names:
    ws = wb.create_sheet(title=name)
    sheets[name] = ws

# Tab colors
tab_colors = {
    "HOME":                    "1B4F2A",
    "MY PROFILE":              "1F3864",
    "POSITION MATCHES":        "1F3864",
    "COMMUTE ANALYSIS":        "1F3864",
    "CAREER TIMELINE":         "1F3864",
    "COUNSELING SHEET":        "1F3864",
    "ADMIN - POSITIONS":       "C00000",
    "ADMIN - CITY DISTANCES":  "C00000",
    "ADMIN - PME AND SCHOOLS": "C00000",
    "ADMIN - MOS TRANSITIONS": "C00000",
    "ADMIN - PROMOTION TIMELINE": "C00000",
    "ADMIN - CONFIG":          "C00000",
    "CALC ENGINE":             "7F7F7F",
}
for name, color in tab_colors.items():
    sheets[name].sheet_properties.tabColor = color

sheets["CALC ENGINE"].sheet_state = "hidden"

print("Sheets created.")

# ═══════════════════════════════════════════════════════════════════════════════
# HOME SHEET
# ═══════════════════════════════════════════════════════════════════════════════
ws = sheets["HOME"]
ws.sheet_view.showGridLines = False
ws.row_dimensions[1].height = 10
ws.row_dimensions[2].height = 50
ws.row_dimensions[3].height = 30

# Title banner
ws.merge_cells("B2:L2")
c = ws["B2"]
c.value = "Montana Army National Guard\nCareer Planning & Mentorship Tool"
c.font = Font(name="Calibri", size=22, bold=True, color=WHITE)
c.fill = fill(ARMY_GREEN)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.merge_cells("B3:L3")
c = ws["B3"]
c.value = "Version 1.0  |  DEMO DATA — Replace admin sheets with real MTARNG data before operational use"
c.font = Font(name="Calibri", size=10, italic=True, color=WHITE)
c.fill = fill(ARMY_BLACK)
c.alignment = center()

# Navigation cards
nav_items = [
    ("MY PROFILE",       "Start here. Enter your rank, MOS,\nunit, home city, and career goals.",    HEADER_BLUE,  "B"),
    ("POSITION MATCHES", "See all MTARNG positions scored\nagainst your profile.",                    ARMY_GREEN,   "D"),
    ("COMMUTE ANALYSIS", "View drive time and distance\nfrom your home city to each position.",       "2E4057",     "F"),
    ("CAREER TIMELINE",  "See a suggested 10–20 year career\npathway toward your goal.",              "5C4033",     "H"),
    ("COUNSELING SHEET", "Print a 1-page counseling summary\nfor your senior rater meeting.",         "6D3B47",     "J"),
]

ws.row_dimensions[5].height = 20
ws.row_dimensions[6].height = 80
ws.row_dimensions[7].height = 30

for title, desc, color, col in nav_items:
    ws.merge_cells(f"{col}5:{col}5")
    ws.merge_cells(f"{col}6:{col}6")
    ws.merge_cells(f"{col}7:{col}7")

    c = ws[f"{col}5"]
    c.value = title
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = fill(color)
    c.alignment = center()

    c = ws[f"{col}6"]
    c.value = desc
    c.font = Font(name="Calibri", size=9, color=WHITE)
    c.fill = fill(color)
    c.alignment = center()

    c = ws[f"{col}7"]
    c.value = f"→ Go to {title}"
    c.font = Font(name="Calibri", size=9, bold=True, color=color)
    c.fill = fill(LIGHT_GRAY)
    c.alignment = center()

ws.column_dimensions["A"].width = 2
for col in ["B","C","D","E","F","G","H","I","J","K","L"]:
    ws.column_dimensions[col].width = 18

# Admin notice
ws.merge_cells("B9:L9")
c = ws["B9"]
c.value = "⚠  ADMIN SHEETS (red tabs): ADMIN - POSITIONS | ADMIN - CITY DISTANCES | ADMIN - PME AND SCHOOLS | ADMIN - MOS TRANSITIONS | ADMIN - PROMOTION TIMELINE"
c.font = Font(name="Calibri", size=9, bold=True, color="C00000")
c.fill = fill("FFF2CC")
c.alignment = center()

# Instructions
instructions = [
    ("STEP 1", "Open MY PROFILE and fill in your information from top to bottom. Use the dropdown lists provided."),
    ("STEP 2", "Navigate to POSITION MATCHES to see every MTARNG position scored and ranked for your profile."),
    ("STEP 3", "Open COMMUTE ANALYSIS to understand the drive time and distance impact of each opportunity."),
    ("STEP 4", "Review CAREER TIMELINE for a suggested multi-year pathway toward your stated goal."),
    ("STEP 5", "Print COUNSELING SHEET and bring it to your senior rater, mentor, or retention NCO."),
    ("FOR ADMINS", "Update the red-tab ADMIN sheets with real MTARNG data. Soldier-facing outputs update automatically."),
]

ws.row_dimensions[11].height = 18
c = ws["B11"]
c.value = "HOW TO USE THIS TOOL"
c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
c.fill = fill(ARMY_GREEN)
c.alignment = center()
ws.merge_cells("B11:L11")

for i, (step, text) in enumerate(instructions, 12):
    ws.row_dimensions[i].height = 22
    c = ws.cell(row=i, column=2, value=step)
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill = fill(ARMY_TAN if i % 2 == 0 else HEADER_BLUE)
    c.alignment = center()

    ws.merge_cells(f"C{i}:L{i}")
    c = ws.cell(row=i, column=3, value=text)
    c.font = body_font(size=9)
    c.fill = fill(LIGHT_BLUE if i % 2 == 0 else LIGHT_GRAY)
    c.alignment = left()

print("HOME sheet done.")

# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN - POSITIONS
# ═══════════════════════════════════════════════════════════════════════════════
ws = sheets["ADMIN - POSITIONS"]
ws.sheet_view.showGridLines = True

# Banner
ws.merge_cells("A1:R1")
c = ws["A1"]
c.value = "⚠ ADMIN SHEET — POSITION REFERENCE DATABASE | Replace [DEMO] rows with real MTARNG UMR/MTOE data"
c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
c.fill = fill("C00000")
c.alignment = center()
ws.row_dimensions[1].height = 20

headers = [
    "Row ID", "Unit Name", "UIC", "City / Location", "Duty Title",
    "Grade", "Career Category", "MOS / AOC", "Branch / FA",
    "Position Type", "Command or KD", "Status Type",
    "Vacancy Status", "Proj Vacancy Date", "Secondary MOS OK",
    "Match Weight Override", "Notes", "DEMO FLAG"
]
write_header(ws, 2, headers, bg=HEADER_BLUE)
ws.row_dimensions[2].height = 30
ws.freeze_panes = "A3"

# Column widths
col_w = {
    "A":6, "B":30, "C":10, "D":16, "E":28,
    "F":8, "G":12, "H":10, "I":10, "J":14,
    "K":12, "L":10, "M":12, "N":16, "O":16,
    "P":14, "Q":35, "R":10
}
set_col_widths(ws, col_w)

# Demo data — 60 positions across Montana
demo_positions = [
    # (RowID, Unit, UIC, City, Title, Grade, Cat, MOS, Branch, PosType, KD, Status, Vacancy, ProjVac, SecMOS, Weight, Notes, Demo)
    # ── BILLINGS ───────────────────────────────────────────────────────────────
    (1,  "1-189th IN BN, A Co",         "W4MMB3","Billings",    "Rifle Team Leader",           "E5","Enlisted","11B","IN","Leadership","No","M-Day","Filled","","","","Squad/team ldr pos","[DEMO]"),
    (2,  "1-189th IN BN, A Co",         "W4MMB3","Billings",    "Squad Leader",                "E6","Enlisted","11B","IN","Leadership","Yes","M-Day","Vacant","2026-09-01","","","KD for promotable SSG","[DEMO]"),
    (3,  "1-189th IN BN, HHC",          "W4MMB4","Billings",    "Operations SGT",              "E7","Enlisted","11B","IN","Staff","No","M-Day","Filled","","","","Batt ops shop","[DEMO]"),
    (4,  "1-189th IN BN, HHC",          "W4MMB4","Billings",    "Readiness NCO",               "E7","Enlisted","92Y","AG","Staff","No","AGR","Filled","","","","Full-time readiness","[DEMO]"),
    (5,  "1-189th IN BN, HHC",          "W4MMB4","Billings",    "S4 NCOIC",                    "E7","Enlisted","92Y","LG","Staff","No","M-Day","Vacant","","","","Log/supply focus","[DEMO]"),
    (6,  "1-189th IN BN, HHC",          "W4MMB4","Billings",    "1SG, A Co",                   "E8","Enlisted","11B","IN","Command","Yes","M-Day","Filled","","","","Senior Enlisted Leader","[DEMO]"),
    (7,  "1-189th IN BN",               "W4MMB4","Billings",    "CSM, 1-189th IN BN",          "E9","Enlisted","11B","IN","Command","Yes","M-Day","Filled","2027-06-01","","","BN CSM","[DEMO]"),
    (8,  "1-189th IN BN",               "W4MMB4","Billings",    "S3, 1-189th IN BN",           "O3","Officer","11A","IN","Staff","Yes","M-Day","Filled","","","","KD staff tour","[DEMO]"),
    (9,  "1-189th IN BN",               "W4MMB4","Billings",    "CDR, A Co, 1-189th IN BN",    "O3","Officer","11A","IN","Command","Yes","M-Day","Vacant","2026-07-01","","","Company command window","[DEMO]"),
    (10, "1-189th IN BN",               "W4MMB4","Billings",    "XO, 1-189th IN BN",           "O4","Officer","11A","IN","Staff","Yes","M-Day","Filled","","","","BN XO","[DEMO]"),
    (11, "1-189th IN BN",               "W4MMB4","Billings",    "CDR, 1-189th IN BN",          "O5","Officer","11A","IN","Command","Yes","M-Day","Filled","2028-01-01","","","BN Command","[DEMO]"),
    (12, "163d Cavalry, HHT",           "W4MM92","Billings",    "Supply SGT",                  "E6","Enlisted","92Y","LG","Technical","No","M-Day","Filled","","","","Unit supply","[DEMO]"),
    (13, "163d Cavalry, HHT",           "W4MM92","Billings",    "Property Book NCO",           "E7","Enlisted","92Y","LG","Staff","No","AGR","Vacant","2026-10-01","","","AGR PB NCO","[DEMO]"),
    (14, "163d Cavalry, HHT",           "W4MM92","Billings",    "Warrant Officer — Property",  "W2","Warrant","920A","LG","Technical","No","M-Day","Filled","","","","Property Accounting Tech","[DEMO]"),
    (15, "163d Cavalry, HHT",           "W4MM92","Billings",    "Troop CDR, HHT",              "O3","Officer","19A","AR","Command","Yes","M-Day","Filled","","","","","[DEMO]"),
    # ── HELENA ────────────────────────────────────────────────────────────────
    (16, "MTARNG G1 / USPFO",           "W4M001","Helena",      "Personnel SGT",               "E5","Enlisted","42A","AG","Staff","No","AGR","Filled","","","","State HQ HR","[DEMO]"),
    (17, "MTARNG G1 / USPFO",           "W4M001","Helena",      "HR NCOIC",                    "E7","Enlisted","42A","AG","Staff","No","AGR","Filled","","","","","[DEMO]"),
    (18, "MTARNG G3",                   "W4M001","Helena",      "Training NCO",                "E7","Enlisted","11B","IN","Staff","No","AGR","Vacant","2027-03-01","","","State G3 shop","[DEMO]"),
    (19, "MTARNG G4",                   "W4M001","Helena",      "Logistics SGT",               "E6","Enlisted","92A","LG","Staff","No","AGR","Filled","","","","State log shop","[DEMO]"),
    (20, "MTARNG G4",                   "W4M001","Helena",      "G4 NCOIC",                    "E8","Enlisted","92A","LG","Staff","No","AGR","Filled","2027-06-01","","","","[DEMO]"),
    (21, "MTARNG HQ",                   "W4M001","Helena",      "State Command Sergeant Major", "E9","Enlisted","00Z","IN","Command","Yes","AGR","Filled","","","","State CSM","[DEMO]"),
    (22, "MTARNG HQ",                   "W4M001","Helena",      "G1, MTARNG",                  "O5","Officer","42H","AG","Staff","Yes","AGR","Filled","","","","","[DEMO]"),
    (23, "MTARNG HQ",                   "W4M001","Helena",      "G3, MTARNG",                  "O5","Officer","11A","IN","Staff","Yes","AGR","Filled","2027-09-01","","","","[DEMO]"),
    (24, "MTARNG HQ",                   "W4M001","Helena",      "The Adjutant General",        "O8","Officer","00A","IN","Command","Yes","AGR","Filled","","","","TAG","[DEMO]"),
    (25, "240th Engineer Co",           "W4MMEC","Helena",      "Combat Engineer",             "E4","Enlisted","12B","EN","Technical","No","M-Day","Vacant","","","","","[DEMO]"),
    (26, "240th Engineer Co",           "W4MMEC","Helena",      "Equipment Operator",          "E5","Enlisted","12B","EN","Technical","No","M-Day","Filled","","","","","[DEMO]"),
    (27, "240th Engineer Co",           "W4MMEC","Helena",      "Squad Leader",                "E6","Enlisted","12B","EN","Leadership","Yes","M-Day","Vacant","2026-10-01","","","","[DEMO]"),
    (28, "240th Engineer Co",           "W4MMEC","Helena",      "Platoon Sergeant",            "E7","Enlisted","12B","EN","Leadership","Yes","M-Day","Filled","","","","","[DEMO]"),
    (29, "240th Engineer Co",           "W4MMEC","Helena",      "1SG, 240th Eng Co",           "E8","Enlisted","12B","EN","Command","Yes","M-Day","Filled","2028-01-01","","","","[DEMO]"),
    (30, "240th Engineer Co",           "W4MMEC","Helena",      "CDR, 240th Engineer Co",      "O3","Officer","12A","EN","Command","Yes","M-Day","Vacant","2026-07-01","","","Company command","[DEMO]"),
    # ── GREAT FALLS ───────────────────────────────────────────────────────────
    (31, "1-163d IN BN, B Co",          "W4MMBF","Great Falls", "Rifle Team Leader",           "E5","Enlisted","11B","IN","Leadership","No","M-Day","Filled","","","","","[DEMO]"),
    (32, "1-163d IN BN, B Co",          "W4MMBF","Great Falls", "Squad Leader",                "E6","Enlisted","11B","IN","Leadership","Yes","M-Day","Filled","","","","","[DEMO]"),
    (33, "1-163d IN BN, B Co",          "W4MMBF","Great Falls", "Weapons Squad Leader",        "E6","Enlisted","11B","IN","Leadership","Yes","M-Day","Vacant","2026-09-01","","","Weapons platoon","[DEMO]"),
    (34, "1-163d IN BN, HHC",           "W4MMBG","Great Falls", "Comms SGT",                   "E6","Enlisted","25U","SC","Technical","No","M-Day","Vacant","","","","","[DEMO]"),
    (35, "1-163d IN BN, HHC",           "W4MMBG","Great Falls", "S6 NCOIC",                    "E7","Enlisted","25U","SC","Staff","No","M-Day","Filled","","","","","[DEMO]"),
    (36, "1-163d IN BN, HHC",           "W4MMBG","Great Falls", "Network Tech WO",             "W2","Warrant","255A","SC","Technical","No","M-Day","Filled","","","","","[DEMO]"),
    (37, "1-163d IN BN, HHC",           "W4MMBG","Great Falls", "Intel Officer",               "O2","Officer","35D","MI","Staff","No","M-Day","Vacant","","","","HUMINT/GEOINT","[DEMO]"),
    (38, "1-163d IN BN, HHC",           "W4MMBG","Great Falls", "S2, 1-163d IN BN",            "O3","Officer","35D","MI","Staff","Yes","M-Day","Filled","","","","","[DEMO]"),
    (39, "1-163d IN BN",                "W4MMBG","Great Falls", "CDR, B Co",                   "O3","Officer","11A","IN","Command","Yes","M-Day","Filled","2027-07-01","","","","[DEMO]"),
    (40, "1-163d IN BN",                "W4MMBG","Great Falls", "XO, 1-163d IN BN",            "O4","Officer","11A","IN","Staff","Yes","M-Day","Filled","","","","","[DEMO]"),
    # ── MISSOULA ──────────────────────────────────────────────────────────────
    (41, "495th Combat Sust Spt Co",    "W4MMMS","Missoula",    "Motor Transport Operator",    "E4","Enlisted","88M","TC","Technical","No","M-Day","Vacant","","","","","[DEMO]"),
    (42, "495th Combat Sust Spt Co",    "W4MMMS","Missoula",    "Motor SGT",                   "E6","Enlisted","88M","TC","Leadership","No","M-Day","Filled","","","","","[DEMO]"),
    (43, "495th Combat Sust Spt Co",    "W4MMMS","Missoula",    "Supply SGT",                  "E6","Enlisted","92Y","LG","Technical","No","M-Day","Vacant","2026-09-01","","","","[DEMO]"),
    (44, "495th Combat Sust Spt Co",    "W4MMMS","Missoula",    "Orderly Room / HR SGT",       "E6","Enlisted","42A","AG","Staff","No","M-Day","Filled","","","","","[DEMO]"),
    (45, "495th Combat Sust Spt Co",    "W4MMMS","Missoula",    "1SG, 495th CSSC",             "E8","Enlisted","92Y","LG","Command","Yes","M-Day","Filled","","","","","[DEMO]"),
    (46, "495th Combat Sust Spt Co",    "W4MMMS","Missoula",    "CDR, 495th CSSC",             "O3","Officer","92A","LG","Command","Yes","M-Day","Vacant","2026-10-01","","","","[DEMO]"),
    # ── BOZEMAN ───────────────────────────────────────────────────────────────
    (47, "163d CAV, A Troop",           "W4MMBT","Bozeman",     "Scout / 19D",                 "E4","Enlisted","19D","AR","Technical","No","M-Day","Vacant","","","","","[DEMO]"),
    (48, "163d CAV, A Troop",           "W4MMBT","Bozeman",     "Scout SGT",                   "E5","Enlisted","19D","AR","Leadership","No","M-Day","Filled","","","","","[DEMO]"),
    (49, "163d CAV, A Troop",           "W4MMBT","Bozeman",     "Staff SGT Scout",             "E6","Enlisted","19D","AR","Leadership","Yes","M-Day","Vacant","","","","","[DEMO]"),
    (50, "163d CAV, A Troop",           "W4MMBT","Bozeman",     "Medic 68W",                   "E5","Enlisted","68W","MC","Technical","No","M-Day","Filled","","","","","[DEMO]"),
    (51, "163d CAV, A Troop",           "W4MMBT","Bozeman",     "Senior Medic / Medical SGT",  "E7","Enlisted","68W","MC","Technical","No","M-Day","Vacant","2027-01-01","","","","[DEMO]"),
    (52, "163d CAV, A Troop",           "W4MMBT","Bozeman",     "Troop XO",                    "O2","Officer","19A","AR","Staff","No","M-Day","Filled","","","","","[DEMO]"),
    (53, "163d CAV, A Troop",           "W4MMBT","Bozeman",     "Troop CDR, A Trp",            "O3","Officer","19A","AR","Command","Yes","M-Day","Filled","2027-06-01","","","","[DEMO]"),
    # ── KALISPELL ─────────────────────────────────────────────────────────────
    (54, "224th Eng Co",                "W4MMKL","Kalispell",   "Combat Engineer",             "E4","Enlisted","12B","EN","Technical","No","M-Day","Vacant","","","","","[DEMO]"),
    (55, "224th Eng Co",                "W4MMKL","Kalispell",   "Squad Leader",                "E6","Enlisted","12B","EN","Leadership","Yes","M-Day","Filled","","","","","[DEMO]"),
    (56, "224th Eng Co",                "W4MMKL","Kalispell",   "Platoon Sergeant",            "E7","Enlisted","12B","EN","Leadership","Yes","M-Day","Vacant","2026-11-01","","","","[DEMO]"),
    (57, "224th Eng Co",                "W4MMKL","Kalispell",   "1SG, 224th Eng Co",           "E8","Enlisted","12B","EN","Command","Yes","M-Day","Filled","","","","","[DEMO]"),
    (58, "224th Eng Co",                "W4MMKL","Kalispell",   "CDR, 224th Eng Co",           "O3","Officer","12A","EN","Command","Yes","M-Day","Filled","2028-01-01","","","","[DEMO]"),
    # ── MILES CITY ────────────────────────────────────────────────────────────
    (59, "163d CAV, C Troop",           "W4MMMC","Miles City",  "Scout",                       "E4","Enlisted","19D","AR","Technical","No","M-Day","Vacant","","","","","[DEMO]"),
    (60, "163d CAV, C Troop",           "W4MMMC","Miles City",  "Scout SGT",                   "E5","Enlisted","19D","AR","Leadership","No","M-Day","Filled","","","","","[DEMO]"),
    (61, "163d CAV, C Troop",           "W4MMMC","Miles City",  "Supply SGT",                  "E6","Enlisted","92Y","LG","Technical","No","M-Day","Vacant","","","","","[DEMO]"),
    (62, "163d CAV, C Troop",           "W4MMMC","Miles City",  "1SG, C Troop",                "E8","Enlisted","19D","AR","Command","Yes","M-Day","Filled","","","","","[DEMO]"),
    (63, "163d CAV, C Troop",           "W4MMMC","Miles City",  "Troop CDR, C Troop",          "O3","Officer","19A","AR","Command","Yes","M-Day","Filled","2027-08-01","","","","[DEMO]"),
    # ── DILLON ────────────────────────────────────────────────────────────────
    (64, "1-189th IN BN, D Co",         "W4MMDI","Dillon",      "Rifle SGT",                   "E5","Enlisted","11B","IN","Leadership","No","M-Day","Filled","","","","","[DEMO]"),
    (65, "1-189th IN BN, D Co",         "W4MMDI","Dillon",      "Squad Leader",                "E6","Enlisted","11B","IN","Leadership","Yes","M-Day","Vacant","2026-10-01","","","","[DEMO]"),
    (66, "1-189th IN BN, D Co",         "W4MMDI","Dillon",      "1SG, D Co",                   "E8","Enlisted","11B","IN","Command","Yes","M-Day","Filled","","","","","[DEMO]"),
    (67, "1-189th IN BN, D Co",         "W4MMDI","Dillon",      "CDR, D Co",                   "O3","Officer","11A","IN","Command","Yes","M-Day","Filled","2028-01-01","","","","[DEMO]"),
    # ── HAVRE ─────────────────────────────────────────────────────────────────
    (68, "163d CAV, D Troop",           "W4MMHV","Havre",       "Scout",                       "E4","Enlisted","19D","AR","Technical","No","M-Day","Vacant","","","","","[DEMO]"),
    (69, "163d CAV, D Troop",           "W4MMHV","Havre",       "Scout SGT",                   "E5","Enlisted","19D","AR","Leadership","No","M-Day","Filled","","","","","[DEMO]"),
    (70, "163d CAV, D Troop",           "W4MMHV","Havre",       "Supply SGT",                  "E6","Enlisted","92Y","LG","Technical","No","M-Day","Filled","","","","","[DEMO]"),
    (71, "163d CAV, D Troop",           "W4MMHV","Havre",       "1SG, D Troop",                "E8","Enlisted","19D","AR","Command","Yes","M-Day","Filled","2028-06-01","","","","[DEMO]"),
    (72, "163d CAV, D Troop",           "W4MMHV","Havre",       "Troop CDR, D Troop",          "O3","Officer","19A","AR","Command","Yes","M-Day","Filled","","","","","[DEMO]"),
    # ── LEWISTOWN / STATEWIDE ─────────────────────────────────────────────────
    (73, "MTARNG JFHQ Med Det",         "W4MHMD","Helena",      "Physician Assistant PA",      "O3","Officer","67D","MC","Technical","No","AGR","Vacant","","","","","[DEMO]"),
    (74, "MTARNG JFHQ Med Det",         "W4MHMD","Helena",      "Medical SGT",                 "E7","Enlisted","68W","MC","Technical","No","AGR","Filled","","","","","[DEMO]"),
    (75, "MTARNG JFHQ",                 "W4M001","Helena",      "PAO NCO",                     "E6","Enlisted","46Z","PA","Staff","No","AGR","Filled","","","","","[DEMO]"),
    (76, "MTARNG Aviation",             "W4MMAV","Helena",      "UH-60 Pilot",                 "W2","Warrant","153A","AV","Technical","No","M-Day","Vacant","","","","IERW required","[DEMO]"),
    (77, "MTARNG Aviation",             "W4MMAV","Helena",      "UH-60 Pilot / PC",            "W3","Warrant","153D","AV","Technical","No","M-Day","Filled","","","","","[DEMO]"),
    (78, "MTARNG Aviation",             "W4MMAV","Helena",      "Aviation CDR",                "O4","Officer","15A","AV","Command","Yes","M-Day","Filled","2028-01-01","","","","[DEMO]"),
    (79, "MTARNG Recruiting & Ret",     "W4MMRR","Helena",      "Retention NCO",               "E7","Enlisted","79S","AG","Staff","No","AGR","Vacant","","","","Statewide travel req","[DEMO]"),
    (80, "MTARNG Recruiting & Ret",     "W4MMRR","Helena",      "Senior Retention NCO",        "E8","Enlisted","79S","AG","Staff","No","AGR","Filled","","","","","[DEMO]"),
]

row_colors = [WHITE, LIGHT_GRAY]
for i, pos in enumerate(demo_positions):
    r = i + 3
    bg = row_colors[i % 2]
    for j, val in enumerate(pos):
        c = ws.cell(row=r, column=j+1, value=val)
        c.font = Font(name="Calibri", size=9, color=ARMY_BLACK if j < 17 else DEMO_ORANGE,
                      bold=(j == 17))
        c.fill = fill(bg if j < 17 else "FFF0E0")
        c.alignment = left()
        c.border = thin_border()

ws.auto_filter.ref = f"A2:R{2+len(demo_positions)}"
print(f"ADMIN - POSITIONS done ({len(demo_positions)} demo rows).")

# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN - CITY DISTANCES
# ═══════════════════════════════════════════════════════════════════════════════
ws = sheets["ADMIN - CITY DISTANCES"]
ws.sheet_view.showGridLines = True

ws.merge_cells("A1:D1")
c = ws["A1"]
c.value = "⚠ ADMIN SHEET — MONTANA CITY DRIVE TIME & DISTANCE MATRIX | Verify against Google Maps / OSRM before operational use"
c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
c.fill = fill("C00000")
c.alignment = center()
ws.row_dimensions[1].height = 20

headers2 = ["From City", "To City", "Miles (One Way)", "Drive Time (hrs:min)", "Route Notes", "DEMO FLAG"]
write_header(ws, 2, headers2)
ws.row_dimensions[2].height = 25

# Montana city pairs (one-way approximate drive times/distances)
city_pairs = [
    ("Billings","Bozeman",144,"2:10","I-90 W","[DEMO]"),
    ("Billings","Butte",223,"3:10","I-90 W","[DEMO]"),
    ("Billings","Helena",221,"3:05","I-90 W / US-12 W","[DEMO]"),
    ("Billings","Great Falls",225,"3:15","US-87 N","[DEMO]"),
    ("Billings","Missoula",344,"4:45","I-90 W","[DEMO]"),
    ("Billings","Kalispell",449,"6:10","I-90 W / US-93 N","[DEMO]"),
    ("Billings","Miles City",141,"2:05","I-94 E","[DEMO]"),
    ("Billings","Dillon",302,"4:15","I-90 W / I-15 S","[DEMO]"),
    ("Billings","Havre",231,"3:20","US-87 N","[DEMO]"),
    ("Billings","Lewistown",126,"1:55","US-87 N","[DEMO]"),
    ("Billings","Livingston",115,"1:40","I-90 W","[DEMO]"),
    ("Billings","Glendive",213,"3:00","I-94 E","[DEMO]"),
    ("Bozeman","Helena",95,"1:20","I-90 W / US-12 W","[DEMO]"),
    ("Bozeman","Butte",82,"1:10","I-90 W","[DEMO]"),
    ("Bozeman","Missoula",199,"2:45","I-90 W","[DEMO]"),
    ("Bozeman","Great Falls",225,"3:15","US-89 N","[DEMO]"),
    ("Bozeman","Kalispell",303,"4:15","I-90 W / US-93 N","[DEMO]"),
    ("Bozeman","Dillon",157,"2:15","I-90 W / I-15 S","[DEMO]"),
    ("Bozeman","Miles City",285,"4:00","I-90 E","[DEMO]"),
    ("Bozeman","Livingston",26,"0:25","I-90 E","[DEMO]"),
    ("Helena","Great Falls",91,"1:10","I-15 N","[DEMO]"),
    ("Helena","Missoula",113,"1:35","I-90 W","[DEMO]"),
    ("Helena","Butte",65,"0:55","I-15 S","[DEMO]"),
    ("Helena","Kalispell",217,"3:00","US-93 N","[DEMO]"),
    ("Helena","Dillon",124,"1:50","I-15 S","[DEMO]"),
    ("Helena","Havre",183,"2:35","US-89 N","[DEMO]"),
    ("Great Falls","Havre",115,"1:35","US-87 N","[DEMO]"),
    ("Great Falls","Lewistown",105,"1:30","US-87 E","[DEMO]"),
    ("Great Falls","Missoula",198,"2:45","US-89 S / I-90 W","[DEMO]"),
    ("Great Falls","Kalispell",222,"3:00","US-89 N","[DEMO]"),
    ("Missoula","Kalispell",118,"1:40","US-93 N","[DEMO]"),
    ("Missoula","Butte",129,"1:50","I-90 E","[DEMO]"),
    ("Missoula","Dillon",189,"2:40","I-90 E / I-15 S","[DEMO]"),
    ("Kalispell","Whitefish",16,"0:20","US-93 N","[DEMO]"),
    ("Miles City","Glendive",72,"1:00","I-94 E","[DEMO]"),
    ("Miles City","Havre",291,"4:05","US-12 W / US-87 N","[DEMO]"),
    ("Dillon","Butte",63,"0:55","I-15 N","[DEMO]"),
    ("Butte","Anaconda",25,"0:30","MT-1","[DEMO]"),
    ("Havre","Lewistown",127,"1:50","US-87 S","[DEMO]"),
    ("Havre","Glasgow",171,"2:25","US-2 E","[DEMO]"),
    ("Glasgow","Wolf Point",49,"0:40","US-2 E","[DEMO]"),
    ("Great Falls","Glasgow",228,"3:10","US-2 E","[DEMO]"),
]

col_w2 = {"A":16,"B":16,"C":16,"D":16,"E":30,"F":10}
set_col_widths(ws, col_w2)

for i, row_data in enumerate(city_pairs):
    r = i + 3
    bg = WHITE if i % 2 == 0 else LIGHT_GRAY
    for j, val in enumerate(row_data):
        c = ws.cell(row=r, column=j+1, value=val)
        c.font = Font(name="Calibri", size=9, color=ARMY_BLACK if j < 5 else DEMO_ORANGE, bold=(j==5))
        c.fill = fill(bg if j < 5 else "FFF0E0")
        c.alignment = center() if j in (2,3) else left()
        c.border = thin_border()

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:F{2+len(city_pairs)}"
print(f"ADMIN - CITY DISTANCES done ({len(city_pairs)} routes).")

# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN - PME AND SCHOOLS
# ═══════════════════════════════════════════════════════════════════════════════
ws = sheets["ADMIN - PME AND SCHOOLS"]

ws.merge_cells("A1:J1")
c = ws["A1"]
c.value = "⚠ ADMIN SHEET — PME & SCHOOL REQUIREMENTS | Update as AR 350-1 / ARNG policy changes"
c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
c.fill = fill("C00000")
c.alignment = center()
ws.row_dimensions[1].height = 20

pme_headers = [
    "Career Cat", "Grade", "PME / Course", "Course Code",
    "Required or Recommended", "Prerequisite", "Duration",
    "Where Taught", "Notes", "DEMO FLAG"
]
write_header(ws, 2, pme_headers)
ws.row_dimensions[2].height = 25

pme_data = [
    # Enlisted
    ("Enlisted","E4-E5","Basic Leader Course (BLC)","BLC","Required for E5 promotion","SSD1 complete","3 weeks","Regional NCO Academy","Must complete before SPC-SGT promotion board","[DEMO]"),
    ("Enlisted","E5-E6","Advanced Leader Course (ALC)","ALC","Required for E6 promotion","BLC complete","4 weeks","Regional NCO Academy","Selectee attendance within 1 year of board selection","[DEMO]"),
    ("Enlisted","E6-E7","Senior Leader Course (SLC)","SLC","Required for E7 promotion","ALC complete","6 weeks","Sergeant Major Academy (USASMA)","Competitive attendance; centrally managed","[DEMO]"),
    ("Enlisted","E8-E9","Sergeant Major Course (SMC)","SMC","Required for CSM selection","SLC complete","10 months","USASMA Fort Bliss TX","Highly competitive; CSM prerequisite","[DEMO]"),
    ("Enlisted","E5+","Structured Self Development 1 (SSD1)","SSD1","Required for BLC attendance","None","Online","AKO / Army e-Learning","Complete prior to BLC registration","[DEMO]"),
    ("Enlisted","E5+","Structured Self Development 2 (SSD2)","SSD2","Required for ALC attendance","BLC + SSD1","Online","AKO / Army e-Learning","","[DEMO]"),
    # Warrant
    ("Warrant","W1","Warrant Officer Basic Course (WOBC)","WOBC","Required after appointment","CW1 appointment","3–6 months (MOS dependent)","MOS branch school","Complete within 24 months of appointment","[DEMO]"),
    ("Warrant","W2-W3","Warrant Officer Advanced Course (WOAC)","WOAC","Required for CW3 promotion","WOBC complete","3 weeks","MOS branch school","Competitive; centrally managed","[DEMO]"),
    ("Warrant","W3-W4","Warrant Officer Intermediate Level Education (WOILE)","WOILE","Required for CW4 promotion","WOAC complete","3 weeks","Ft Leavenworth or DL","","[DEMO]"),
    ("Warrant","W4-W5","Warrant Officer Senior Service Education (WOSSE)","WOSSE","Required for CW5 selection","WOILE complete","2 weeks","Various","Seminar-based","[DEMO]"),
    # Officer
    ("Officer","O1","Basic Officer Leadership Course (BOLC)","BOLC","Required after commissioning","Commission","8–24 weeks (branch dep)","Branch school","Complete within 12 months of commissioning","[DEMO]"),
    ("Officer","O3","Captain's Career Course (CCC)","CCC","Required for MAJ promotion","BOLC + time","18 weeks","Branch school","Complete prior to MAJ promotion board","[DEMO]"),
    ("Officer","O4-O5","Intermediate Level Education (ILE)","ILE","Required for LTC promotion","CCC + 8 yrs","10 months","Ft Leavenworth (CGSC) or non-resident","Highly competitive resident; non-resident via ATRRS","[DEMO]"),
    ("Officer","O5-O6","Senior Service College (SSC)","SSC","Required for COL/GO","ILE complete","10 months","AWC, ICAF, or equivalent","Very competitive; <15% selected","[DEMO]"),
    ("Officer","O3+","Pre-Command Course (PCC)","PCC","Required before company command","CCC + slated","1 week","Fort Leavenworth","Required for all company commanders","[DEMO]"),
    ("Officer","O4+","Battalion Pre-Command Course (BN PCC)","BN-PCC","Required before BN command","ILE + slated","2 weeks","Fort Leavenworth","Required for LTC command selection","[DEMO]"),
    # MOS-specific
    ("Enlisted","E5-E6","92Y Unit Supply Specialist Advanced","92Y ALC","Required for 92Y E6","BLC + 92Y MOSQ","4 weeks","Quartermasters (Ft Lee)","MOS-specific ALC track","[DEMO]"),
    ("Enlisted","E5-E6","68W Combat Medic Specialist Advanced","68W ALC","Required for 68W E6","BLC + EMT-B","4 weeks","Medical Center of Excellence","EMT-B required for 68W","[DEMO]"),
    ("Enlisted","Any","Recruiting and Retention School","RRS","Required for 79S MOS","SSG+ recommended","4 weeks","Ft Knox","Required for retention NCO qualification","[DEMO]"),
    ("Warrant","W1-W2","Initial Entry Rotary Wing (IERW)","IERW","Required for aviators","WOFT selection","14–18 months","Ft Rucker (Novosel)","For 153A/153D WO aviators","[DEMO]"),
]

pme_col_w = {"A":10,"B":10,"C":35,"D":12,"E":20,"F":18,"G":16,"H":30,"I":40,"J":10}
set_col_widths(ws, pme_col_w)

for i, row_data in enumerate(pme_data):
    r = i + 3
    bg = WHITE if i % 2 == 0 else LIGHT_BLUE
    for j, val in enumerate(row_data):
        c = ws.cell(row=r, column=j+1, value=val)
        c.font = Font(name="Calibri", size=9, color=ARMY_BLACK if j < 9 else DEMO_ORANGE, bold=(j==9))
        c.fill = fill(bg if j < 9 else "FFF0E0")
        c.alignment = left()
        c.border = thin_border()

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:J{2+len(pme_data)}"
print(f"ADMIN - PME AND SCHOOLS done ({len(pme_data)} rows).")

# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN - MOS TRANSITIONS
# ═══════════════════════════════════════════════════════════════════════════════
ws = sheets["ADMIN - MOS TRANSITIONS"]

ws.merge_cells("A1:L1")
c = ws["A1"]
c.value = "⚠ ADMIN SHEET — MOS TRANSITION PATHWAYS | Verify line scores and waivers with G1/Retention before counseling"
c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
c.fill = fill("C00000")
c.alignment = center()
ws.row_dimensions[1].height = 20

mos_headers = [
    "From MOS", "To MOS", "To MOS Title", "Eligible Grades",
    "ASVAB Line Score Req", "MOSQ School", "School Duration",
    "Waiver Available", "Med / Phys Req", "Notes", "Difficulty", "DEMO FLAG"
]
write_header(ws, 2, mos_headers)

mos_data = [
    ("11B","19D","Cavalry Scout","E1-E6","CO:87","OSUT Ft Knox / ANCOC","14–22 wks","Limited","PULHES 111111","Lateral similar track","Moderate","[DEMO]"),
    ("11B","12B","Combat Engineer","E1-E6","CO:87, GT:100","OSUT Ft Leonard Wood","14–18 wks","Limited","PULHES 111111","Popular reclass track","Moderate","[DEMO]"),
    ("11B","25U","Signal Support Spc","E1-E6","EL:93, GT:100","AIT Ft Gordon","20 wks","Yes","Standard","Good broadening option","Moderate","[DEMO]"),
    ("11B","92Y","Unit Supply Spc","E1-E7","CL:90, OF:85","AIT Ft Lee","8 wks","Yes","Standard","Common reclass for senior NCOs","Easy","[DEMO]"),
    ("11B","42A","HR Specialist","E1-E6","CL:90, VE:72","AIT Ft Jackson","10 wks","Yes","Standard","Good AGR pathway","Easy","[DEMO]"),
    ("11B","68W","Combat Medic","E1-E5","ST:101, GT:107","AMEDD AIT + EMT-B","16 wks","No","PULHES 111111, color vision","Physically demanding","Hard","[DEMO]"),
    ("11B","79S","Career Counselor","E6-E8","GT:110","Recruiting & Ret School Ft Knox","4 wks","Yes","Standard","E6 min; good senior career","Moderate","[DEMO]"),
    ("92Y","88M","Motor Transport Op","E1-E6","OF:85","AIT Ft Lee","8 wks","Yes","Class B CDL req","CDL may be waived","Easy","[DEMO]"),
    ("92Y","42A","HR Specialist","E1-E6","CL:90, VE:72","AIT Ft Jackson","10 wks","Yes","Standard","Related admin track","Easy","[DEMO]"),
    ("92Y","92A","Automated Logistical Spc","E1-E6","CL:90, OF:85","AIT Ft Lee","10 wks","Yes","Standard","Promotion pathway for 92Y","Easy","[DEMO]"),
    ("68W","18D","SF Medical SGT","E6+","GT:110, ST:101","SFMS Ft Bragg","57 wks","No","PULHES 111111, Phase I physical","Must volunteer; very selective","Hard","[DEMO]"),
    ("68W","66H","Registered Nurse","O1+","Commission + RN license","AMEDD Officer Basic","Varies","No","Current RN license","Commission pathway","Hard","[DEMO]"),
    ("25U","255A","Network WO","E5-E6 (Warrant)","EL:98, GT:110","WOCC + WOBC Fort Gordon","6 months","No","Standard","WO appointment required","Hard","[DEMO]"),
    ("19D","11A","Infantry Officer","E5-E7 (OCS)","GT:110","OCS + BOLC","18 months","Yes","PULHES 111111","OCS then reclass","Hard","[DEMO]"),
    ("Any","79V","Retention & Trans NCO","E7+","GT:100","Retention School","4 wks","Yes","Standard","State-managed; AGR common","Moderate","[DEMO]"),
    ("12B","12A","Engineer Officer (OCS)","E5-E7","GT:110","OCS + BOLC","18 months","Yes","Standard","Common pathway for senior 12B","Hard","[DEMO]"),
    ("Any","153A","Warrant Aviator","E3-E5 or civilian","GT:110, SIFT score","WOFT + IERW","18–24 months","No","Class 1 flight physical","High demand; very selective","Hard","[DEMO]"),
    ("42A","36A","Financial Manager (Officer)","E4-E6","GT:110","OCS + BOLC FI","18 months","Yes","Standard","Finance officer pathway","Hard","[DEMO]"),
]

mos_col_w = {"A":9,"B":9,"C":25,"D":14,"E":18,"F":28,"G":14,"H":12,"I":22,"J":42,"K":10,"L":10}
set_col_widths(ws, mos_col_w)

for i, row_data in enumerate(mos_data):
    r = i + 3
    # Color difficulty
    diff = row_data[10]
    diff_bg = {"Easy": LIGHT_GREEN, "Moderate": LIGHT_YELLOW, "Hard": LIGHT_RED}.get(diff, WHITE)
    bg = WHITE if i % 2 == 0 else LIGHT_GRAY
    for j, val in enumerate(row_data):
        c = ws.cell(row=r, column=j+1, value=val)
        cell_bg = diff_bg if j == 10 else (bg if j < 11 else "FFF0E0")
        c.font = Font(name="Calibri", size=9,
                      color=ARMY_BLACK if j < 11 else DEMO_ORANGE,
                      bold=(j == 11 or j == 10))
        c.fill = fill(cell_bg)
        c.alignment = left()
        c.border = thin_border()

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:L{2+len(mos_data)}"
print(f"ADMIN - MOS TRANSITIONS done ({len(mos_data)} rows).")

# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN - PROMOTION TIMELINE
# ═══════════════════════════════════════════════════════════════════════════════
ws = sheets["ADMIN - PROMOTION TIMELINE"]

ws.merge_cells("A1:J1")
c = ws["A1"]
c.value = "⚠ ADMIN SHEET — PROMOTION TIMELINE REFERENCE | Verify TIG/TIS against current ARNG selection board criteria"
c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
c.fill = fill("C00000")
c.alignment = center()
ws.row_dimensions[1].height = 20

promo_headers = [
    "Career Category", "Promote To", "Min TIG (months)", "Typical TIG (months)",
    "Min TIS (years)", "PME Required Before Board",
    "Board Type", "Promotion Rate (est)", "Notes", "DEMO FLAG"
]
write_header(ws, 2, promo_headers)

promo_data = [
    # Enlisted
    ("Enlisted","SGT (E5)","12","24","2","BLC enrolled or complete","Unit","~85%","SPC(P) required; most common bottleneck is BLC seat","[DEMO]"),
    ("Enlisted","SSG (E6)","18","36","5","BLC complete; ALC enrolled","State/Unit","~40–55%","Competitive; leadership position often required","[DEMO]"),
    ("Enlisted","SFC (E7)","18","48","8","ALC complete; SLC eligible","DA Centralized","~25–35%","Highly competitive; key developmental required","[DEMO]"),
    ("Enlisted","MSG (E8)","12","60","10","SLC complete","DA Centralized","~15–20%","Senior leader competition; limited MTARNG positions","[DEMO]"),
    ("Enlisted","SGM (E9)","12","72","12","SLC complete","DA Centralized","~5–8%","Limited; SMC for CSM path","[DEMO]"),
    ("Enlisted","1SG (E8 variant)","N/A","N/A","N/A","SLC complete","Command Designation","N/A","1SG is a position, not a separate grade; still MSG","[DEMO]"),
    ("Enlisted","CSM (E9 variant)","N/A","N/A","N/A","SMC","CSM Selection Board","Very competitive","CSM selection separate from SGM board","[DEMO]"),
    # Warrant
    ("Warrant","CW2","18","24","2","WOBC complete","DA","~90%","Nearly automatic with WOBC completion","[DEMO]"),
    ("Warrant","CW3","24","36","5","WOAC complete","DA","~70%","WOAC completion key","[DEMO]"),
    ("Warrant","CW4","24","48","8","WOILE complete","DA","~50%","Selective above CW3","[DEMO]"),
    ("Warrant","CW5","18","72","12","WOSSE complete","DA","~15–20%","Very limited; senior WO leadership","[DEMO]"),
    # Officer
    ("Officer","1LT (O2)","18","18","1.5","N/A","Time-based","~98%","Essentially automatic","[DEMO]"),
    ("Officer","CPT (O3)","24","36","4","BOLC complete","DA","~85–90%","CCC should be complete/enrolled","[DEMO]"),
    ("Officer","MAJ (O4)","36","60","8","CCC complete","DA","~70–80%","ILE completion strongly preferred","[DEMO]"),
    ("Officer","LTC (O5)","24","84","14","ILE complete","DA","~50–60%","Command and KD assignments critical","[DEMO]"),
    ("Officer","COL (O6)","36","120","18","ILE; SSC preferred","DA","~20–30%","SSC near mandatory; very competitive","[DEMO]"),
]

promo_col_w = {"A":13,"B":18,"C":14,"D":16,"E":12,"F":24,"G":14,"H":16,"I":45,"J":10}
set_col_widths(ws, promo_col_w)

grade_colors = {
    "SGT (E5)": "E2EFDA", "SSG (E6)": "DDEBF7", "SFC (E7)": "FCE4D6",
    "MSG (E8)": "F4CCCC", "SGM (E9)": "EAD1DC", "1SG (E8 variant)": "F4CCCC",
    "CSM (E9 variant)": "EAD1DC",
    "CW2": "FFF2CC", "CW3": "FFF2CC", "CW4": "FFF2CC", "CW5": "FFF2CC",
    "1LT (O2)": "D9EAD3", "CPT (O3)": "CFE2F3", "MAJ (O4)": "FCE5CD",
    "LTC (O5)": "EA9999", "COL (O6)": "CC0000",
}

for i, row_data in enumerate(promo_data):
    r = i + 3
    row_color = grade_colors.get(row_data[1], WHITE)
    for j, val in enumerate(row_data):
        c = ws.cell(row=r, column=j+1, value=val)
        c.font = Font(name="Calibri", size=9,
                      color=WHITE if row_data[1] == "COL (O6)" and j < 9 else (ARMY_BLACK if j < 9 else DEMO_ORANGE),
                      bold=(j == 9))
        c.fill = fill(row_color if j < 9 else "FFF0E0")
        c.alignment = center() if j in (2,3,4,7) else left()
        c.border = thin_border()

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:J{2+len(promo_data)}"
print(f"ADMIN - PROMOTION TIMELINE done ({len(promo_data)} rows).")

# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN - CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
ws = sheets["ADMIN - CONFIG"]
ws.sheet_view.showGridLines = False

ws.merge_cells("B2:F2")
c = ws["B2"]
c.value = "ADMIN — TOOL CONFIGURATION SETTINGS"
c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
c.fill = fill(ARMY_GREEN)
c.alignment = center()
ws.row_dimensions[2].height = 28

config_items = [
    ("Tool Name",             "MT ARNG Career Planner",     "Display name in counseling sheet header"),
    ("Tool Version",          "1.0",                         "Update when reference data is refreshed"),
    ("Data As Of Date",       "2026-06-01",                  "Date reference data was last validated"),
    ("Responsible Office",    "MTARNG G1 / S1 Section",      "Office responsible for maintaining this tool"),
    ("Contact Email",         "g1@mt.ngb.army.mil",         "Admin contact for tool questions"),
    ("State",                 "Montana",                     ""),
    ("","",""),
    ("SCORING WEIGHTS","",""),
    ("Grade Match Weight",    35,                            "Points available for grade match (0–35)"),
    ("MOS/AOC Match Weight",  30,                            "Points available for MOS match (0–30)"),
    ("Commute Match Weight",  15,                            "Points available for commute fit (0–15)"),
    ("Career Goal Weight",    15,                            "Points available for goal alignment (0–15)"),
    ("Status Pref Weight",    5,                             "Points available for status preference (0–5)"),
    ("","",""),
    ("MATCH LABELS","",""),
    ("Strong Match Threshold",    80,                        "Score >= this = STRONG MATCH"),
    ("Good Match Threshold",      60,                        "Score >= this = GOOD MATCH"),
    ("Moderate Match Threshold",  40,                        "Score >= this = MODERATE / STRETCH"),
    ("Possible Match Threshold",  20,                        "Score >= this = POSSIBLE — GAPS EXIST"),
    ("Below Threshold Label",     "NOT RECOMMENDED NOW",     "Score < Possible threshold"),
    ("","",""),
    ("COMMUTE CATEGORIES","",""),
    ("Short Commute Max (min)",   30,                        "Drive time <= this = Short"),
    ("Medium Commute Max (min)", 90,                         "Drive time <= this = Medium"),
    ("Long Commute Max (min)",   150,                        "Drive time <= this = Long"),
    ("Very Long Label",           "Very Long / Consider Relocating",""),
]

ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 32
ws.column_dimensions["D"].width = 45

for i, (label, value, note) in enumerate(config_items, 4):
    ws.row_dimensions[i].height = 18
    if not label:
        continue
    if label in ("SCORING WEIGHTS","MATCH LABELS","COMMUTE CATEGORIES"):
        c = ws.cell(row=i, column=2, value=label)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = fill(HEADER_BLUE)
        c.alignment = left()
        ws.merge_cells(f"B{i}:D{i}")
        continue
    c = ws.cell(row=i, column=2, value=label)
    c.font = body_font(bold=True)
    c.fill = fill(LIGHT_GRAY)
    c.alignment = left()
    c.border = thin_border()

    c = ws.cell(row=i, column=3, value=value)
    c.font = Font(name="Calibri", size=10, color="1F3864")
    c.fill = fill(WHITE)
    c.alignment = left()
    c.border = thin_border()

    # Named range-compatible cell reference hint
    c = ws.cell(row=i, column=4, value=note)
    c.font = Font(name="Calibri", size=9, color="595959", italic=True)
    c.fill = fill(LIGHT_GRAY)
    c.alignment = left()
    c.border = thin_border()

print("ADMIN - CONFIG done.")

# Save Part 1
wb.save("/home/user/S1-AI-Assistsnt-/MT_ARNG_Career_Planner.xlsx")
print("\nPart 1 saved: MT_ARNG_Career_Planner.xlsx")
print("Sheets complete:", [s.title for s in wb.worksheets])
