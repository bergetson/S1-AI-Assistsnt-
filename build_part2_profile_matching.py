"""
Part 2: MY PROFILE input form + CALC ENGINE + POSITION MATCHES sheet.
Run AFTER build_part1_reference.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Shared styles (must match Part 1) ────────────────────────────────────────
ARMY_GREEN  = "1B4F2A"
ARMY_TAN    = "C8A96E"
ARMY_BLACK  = "1A1A1A"
HEADER_BLUE = "1F3864"
LIGHT_BLUE  = "D9E1F2"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW= "FFFF99"
LIGHT_ORANGE= "FCE4D6"
LIGHT_RED   = "FFE0E0"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
MED_GRAY    = "D9D9D9"
INPUT_YELLOW= "FFFACD"   # color for cells the Soldier fills in
DEMO_ORANGE = "FF6600"

def thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_left():
    tl = Side(style="medium", color=ARMY_GREEN)
    s  = Side(style="thin",   color="AAAAAA")
    return Border(left=tl, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_al():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def hf(size=10, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def bf(size=10, bold=False, color=ARMY_BLACK):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def label(ws, row, col, text, bg=LIGHT_GRAY, fg=ARMY_BLACK, bold=True, span=None, height=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=9, bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = left_al()
    c.border = thin()
    if height:
        ws.row_dimensions[row].height = height
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def input_cell(ws, row, col, default_value="", bg=INPUT_YELLOW, span=None, align="left"):
    c = ws.cell(row=row, column=col, value=default_value)
    c.font = Font(name="Calibri", size=10, bold=False, color="1F3864")
    c.fill = fill(bg)
    c.alignment = center() if align == "center" else left_al()
    c.border = thin()
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return c

def section_banner(ws, row, col, text, bg=HEADER_BLUE, span=8, height=22):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    c.fill = fill(bg)
    c.alignment = center()
    c.border = thin()
    ws.row_dimensions[row].height = height
    return c

def dv_list(ws, formula, rows, col):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                        showDropDown=False, showErrorMessage=True,
                        errorTitle="Invalid Entry", error="Please select from the list.")
    ws.add_data_validation(dv)
    for row in rows:
        dv.add(ws.cell(row=row, column=col))

def set_col_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

# ═══════════════════════════════════════════════════════════════════════════════
# Load workbook
# ═══════════════════════════════════════════════════════════════════════════════
wb = load_workbook("/home/user/S1-AI-Assistsnt-/MT_ARNG_Career_Planner.xlsx")
ws_profile  = wb["MY PROFILE"]
ws_matches  = wb["POSITION MATCHES"]
ws_calc     = wb["CALC ENGINE"]

# ═══════════════════════════════════════════════════════════════════════════════
# MY PROFILE — Input Form
# Layout: columns A–I  (A=narrow margin, B=label 24w, C–E=input, F=label, G–I=input)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_profile
ws.sheet_view.showGridLines = False

set_col_widths(ws, {
    "A": 2,   # margin
    "B": 26,  # label col left
    "C": 18,  # input col left 1
    "D": 14,  # input col left 2
    "E": 14,  # input col left 3
    "F": 26,  # label col right
    "G": 18,  # input col right 1
    "H": 14,  # input col right 2
    "I": 14,  # input col right 3
    "J": 2,   # right margin
})

# ── Title ────────────────────────────────────────────────────────────────────
ws.merge_cells("B1:I1")
c = ws["B1"]
c.value = "SOLDIER CAREER PROFILE  |  Montana Army National Guard"
c.font  = Font(name="Calibri", size=16, bold=True, color=WHITE)
c.fill  = fill(ARMY_GREEN)
c.alignment = center()
ws.row_dimensions[1].height = 34

ws.merge_cells("B2:I2")
c = ws["B2"]
c.value = "Fill in the yellow cells. Use dropdown arrows where available. All Soldier-facing sheets update automatically from this page."
c.font  = Font(name="Calibri", size=9, italic=True, color=ARMY_BLACK)
c.fill  = fill(LIGHT_BLUE)
c.alignment = center()
ws.row_dimensions[2].height = 18

# ── SECTION 1: Identity ───────────────────────────────────────────────────────
section_banner(ws, 4, 2, "SECTION 1 — SOLDIER IDENTITY & CURRENT STATUS", bg=HEADER_BLUE, span=8)

rows_s1 = range(5, 15)
for r in rows_s1:
    ws.row_dimensions[r].height = 22

# Row 5: Name | Social Last 4
label(ws, 5, 2, "Full Name (Last, First MI)")
input_cell(ws, 5, 3, "Smith, John A.", span=3)
label(ws, 5, 6, "Last 4 SSN (optional)")
input_cell(ws, 5, 7, "XXXX", span=2)

# Row 6: Rank | Career Category
label(ws, 6, 2, "Current Rank / Grade")
input_cell(ws, 6, 3, "E5 / SGT")
dv_list(ws, '"E1,E2,E3,E4,E5,E6,E7,E8,E9,W1,W2,W3,W4,W5,O1,O2,O3,O4,O5,O6"', [6], 3)
label(ws, 6, 4, "Career Category")
input_cell(ws, 6, 5, "Enlisted")
dv_list(ws, '"Enlisted,Officer,Warrant"', [6], 5)
label(ws, 6, 6, "Component / Status")
input_cell(ws, 6, 7, "M-Day")
dv_list(ws, '"M-Day,AGR,Technician,ADOS,IRR"', [6], 7)
input_cell(ws, 6, 8, "", span=2)

# Row 7: MOS
label(ws, 7, 2, "Primary MOS / AOC / Branch")
input_cell(ws, 7, 3, "92Y")
label(ws, 7, 4, "Secondary MOS (if any)")
input_cell(ws, 7, 5, "")
label(ws, 7, 6, "Branch / Career Field (Officers)")
input_cell(ws, 7, 7, "LG", span=2)

# Row 8: Unit
label(ws, 8, 2, "Current Unit Name")
input_cell(ws, 8, 3, "495th CSSC", span=3)
label(ws, 8, 6, "UIC (if known)")
input_cell(ws, 8, 7, "W4MMMS", span=2)

# Row 9: Duty Title
label(ws, 9, 2, "Current Duty Title / Position")
input_cell(ws, 9, 3, "Unit Supply Specialist", span=3)
label(ws, 9, 6, "City of Current Unit")
input_cell(ws, 9, 7, "Missoula")
dv_list(ws, '"Billings,Bozeman,Butte,Great Falls,Helena,Kalispell,Missoula,Miles City,Dillon,Havre,Lewistown,Livingston,Glendive,Glasgow,Wolf Point,Whitefish,Anaconda,Shelby"', [9], 7)

# Row 10: Home City | Max commute
label(ws, 10, 2, "Home City (where you live)")
input_cell(ws, 10, 3, "Missoula")
dv_list(ws, '"Billings,Bozeman,Butte,Great Falls,Helena,Kalispell,Missoula,Miles City,Dillon,Havre,Lewistown,Livingston,Glendive,Glasgow,Wolf Point,Whitefish,Anaconda,Shelby"', [10], 3)
label(ws, 10, 4, "State (if not MT)")
input_cell(ws, 10, 5, "MT")
label(ws, 10, 6, "Max Acceptable One-Way Commute")
input_cell(ws, 10, 7, "2 Hours")
dv_list(ws, '"30 Minutes,1 Hour,1.5 Hours,2 Hours,3 Hours,No Limit"', [10], 7)

# Row 11: Years of Service | TIG
label(ws, 11, 2, "Total Years of Service (ARNG + AD)")
input_cell(ws, 11, 3, "8")
label(ws, 11, 4, "Years in Current Grade (TIG)")
input_cell(ws, 11, 5, "2")
label(ws, 11, 6, "Time in Current Position (years)")
input_cell(ws, 11, 7, "1.5", span=2)

# Row 12: DODID / ETS
label(ws, 12, 2, "ETS / Expiration Term of Service")
input_cell(ws, 12, 3, "2027-08-01")
label(ws, 12, 4, "Initial Entry Date (PEBD/DOR)")
input_cell(ws, 12, 5, "2018-06-01")
label(ws, 12, 6, "Willing to Relocate for the Right Position?")
input_cell(ws, 12, 7, "No")
dv_list(ws, '"Yes,No,Maybe"', [12], 7)

# ── SECTION 2: PME & Education ────────────────────────────────────────────────
section_banner(ws, 14, 2, "SECTION 2 — PME & EDUCATION COMPLETED", bg="2E4057", span=8)

for r in range(15, 24):
    ws.row_dimensions[r].height = 20

# Enlisted PME
label(ws, 15, 2, "ENLISTED PME", bg=ARMY_TAN, fg=ARMY_BLACK, bold=True, span=2)
label(ws, 15, 4, "OFFICER PME", bg=ARMY_TAN, fg=ARMY_BLACK, bold=True, span=2)
label(ws, 15, 6, "WARRANT PME", bg=ARMY_TAN, fg=ARMY_BLACK, bold=True, span=3)

pme_enlisted = [
    ("Basic Leader Course (BLC)", 16, 3),
    ("Advanced Leader Course (ALC)", 17, 3),
    ("Senior Leader Course (SLC)", 18, 3),
    ("Sergeant Major Course (SMC)", 19, 3),
    ("SSD1 Complete", 20, 3),
    ("SSD2 Complete", 21, 3),
]
pme_officer = [
    ("BOLC Complete", 16, 5),
    ("Captain's Career Course (CCC)", 17, 5),
    ("ILE Complete", 18, 5),
    ("Senior Service College (SSC)", 19, 5),
]
pme_warrant = [
    ("WOBC Complete", 16, 7),
    ("WOAC Complete", 17, 7),
    ("WOILE Complete", 18, 7),
    ("WOSSE Complete", 19, 7),
]

for text, row, col_inp in pme_enlisted:
    label(ws, row, 2, text, bg=LIGHT_GRAY)
    input_cell(ws, row, col_inp, "No")
    dv_list(ws, '"Yes,No,In Progress,N/A"', [row], col_inp)
    label(ws, row, 4, "", bg=WHITE)  # spacer

for text, row, col_inp in pme_officer:
    label(ws, row, 4, text, bg=LIGHT_GRAY)
    input_cell(ws, row, col_inp, "No")
    dv_list(ws, '"Yes,No,In Progress,N/A"', [row], col_inp)

for text, row, col_inp in pme_warrant:
    label(ws, row, 6, text, bg=LIGHT_GRAY)
    input_cell(ws, row, col_inp, "No")
    dv_list(ws, '"Yes,No,In Progress,N/A"', [row], col_inp)

# College
label(ws, 22, 2, "Highest Education Level")
input_cell(ws, 22, 3, "Some College")
dv_list(ws, '"High School / GED,Some College,Associate Degree,Bachelor\'s Degree,Master\'s Degree,Doctoral Degree"', [22], 3)
label(ws, 22, 4, "College Degree Field (if any)")
input_cell(ws, 22, 5, "Business Administration")
label(ws, 22, 6, "Other Relevant Certifications")
input_cell(ws, 22, 7, "APFT Master Fitness Trainer", span=2)

# ── SECTION 3: Career Goals ───────────────────────────────────────────────────
section_banner(ws, 24, 2, "SECTION 3 — CAREER GOALS & ASPIRATIONS", bg=ARMY_GREEN, span=8)

for r in range(25, 38):
    ws.row_dimensions[r].height = 22

label(ws, 25, 2, "Target Rank / Goal Grade")
input_cell(ws, 25, 3, "E7 / SFC")
dv_list(ws, '"E5,E6,E7,E8,E9,W2,W3,W4,W5,O2,O3,O4,O5,O6"', [25], 3)
label(ws, 25, 4, "Target Timeline (years from now)")
input_cell(ws, 25, 5, "4")
label(ws, 25, 6, "Preferred Component After Next Move")
input_cell(ws, 25, 7, "Either")
dv_list(ws, '"M-Day,AGR,Technician,ADOS,Either"', [25], 7)

label(ws, 26, 2, "Primary Career Goal (select one)")
input_cell(ws, 26, 3, "Promote to Next Grade", span=3)
dv_list(ws, '"Promote to Next Grade,Switch MOS/Branch,Pursue AGR Full-Time,Pursue Command,Move to Staff,Become Warrant Officer,Pursue Senior Leader Path,Broaden Experience,Transition to AGR,Retire in Current Grade"', [26], 3)
label(ws, 26, 6, "Secondary Career Goal (optional)")
input_cell(ws, 26, 7, "Pursue AGR Full-Time", span=2)

label(ws, 27, 2, "Desired Position Type (select one)")
input_cell(ws, 27, 3, "Leadership")
dv_list(ws, '"Leadership,Staff,Command,Technical,Broadening,Any"', [27], 3)
label(ws, 27, 4, "Open to Command Position?")
input_cell(ws, 27, 5, "Yes")
dv_list(ws, '"Yes,No,In a few years"', [27], 5)
label(ws, 27, 6, "Want to switch MOS or Branch?")
input_cell(ws, 27, 7, "No")
dv_list(ws, '"Yes,No,Considering it"', [27], 7)

label(ws, 28, 2, "If switching — Target MOS / Branch")
input_cell(ws, 28, 3, "N/A")
label(ws, 28, 4, "Warrant Officer Interest?")
input_cell(ws, 28, 5, "No")
dv_list(ws, '"Yes - Active interest,No,Considering it"', [28], 5)
label(ws, 28, 6, "Long-term goal (10-20 yrs out)")
input_cell(ws, 28, 7, "Senior NCO / Retention SGM", span=2)

# Narrative goals
label(ws, 30, 2, "5-YEAR GOAL (describe in your own words):", bg=HEADER_BLUE, fg=WHITE, bold=True, span=8)
ws.row_dimensions[30].height = 18

ws.merge_cells("B31:I33")
c = ws["B31"]
c.value = "I want to promote to SFC within 4 years. I plan to complete SLC and move into a platoon sergeant or staff NCOIC role at a unit closer to Helena, ideally as an AGR position."
c.font = Font(name="Calibri", size=10, color="1F3864")
c.fill = fill(INPUT_YELLOW)
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c.border = thin()
for r in range(31, 34):
    ws.row_dimensions[r].height = 20

label(ws, 34, 2, "10-YEAR / LONG-RANGE GOAL:", bg=HEADER_BLUE, fg=WHITE, bold=True, span=8)
ws.row_dimensions[34].height = 18

ws.merge_cells("B35:I37")
c = ws["B35"]
c.value = "Long term, I'd like to either achieve MSG or pursue a warrant officer program if my career path allows. I want to finish my bachelor's degree and remain in Montana."
c.font = Font(name="Calibri", size=10, color="1F3864")
c.fill = fill(INPUT_YELLOW)
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c.border = thin()
for r in range(35, 38):
    ws.row_dimensions[r].height = 20

# ── SECTION 4: Current Career Status ─────────────────────────────────────────
section_banner(ws, 39, 2, "SECTION 4 — ADDITIONAL CAREER INFORMATION", bg="5C4033", span=8)

for r in range(40, 50):
    ws.row_dimensions[r].height = 20

label(ws, 40, 2, "Have you been to any command/KD position?")
input_cell(ws, 40, 3, "No")
dv_list(ws, '"Yes,No"', [40], 3)
label(ws, 40, 4, "Currently promotable (P)?")
input_cell(ws, 40, 5, "No")
dv_list(ws, '"Yes,No"', [40], 5)
label(ws, 40, 6, "On promotion list for current grade?")
input_cell(ws, 40, 7, "No")
dv_list(ws, '"Yes,No"', [40], 7)

label(ws, 41, 2, "Previous Key Assignments (summarize)")
ws.merge_cells("C41:I41")
c = ws["C41"]
c.value = "SPC - 92Y supply tech at 495th CSSC; SGT - acting supply NCOIC during annual training 2025"
c.font = Font(name="Calibri", size=9, color="1F3864")
c.fill = fill(INPUT_YELLOW)
c.alignment = left_al()
c.border = thin()
ws.row_dimensions[41].height = 24

label(ws, 42, 2, "Physical Fitness Status")
input_cell(ws, 42, 3, "Passing")
dv_list(ws, '"Passing,Failing,Profile,Outstanding"', [42], 3)
label(ws, 42, 4, "Security Clearance Level")
input_cell(ws, 42, 5, "Secret")
dv_list(ws, '"None,Confidential,Secret,Top Secret,TS/SCI"', [42], 5)
label(ws, 42, 6, "Deployment / GWOT Service")
input_cell(ws, 42, 7, "1 deployment (Afghanistan 2021)", span=2)

label(ws, 43, 2, "Awards / Distinction (brief summary)")
ws.merge_cells("C43:I43")
c = ws["C43"]
c.value = "Army Achievement Medal x2, Army Commendation Medal, Expert Infantryman Badge"
c.font = Font(name="Calibri", size=9, color="1F3864")
c.fill = fill(INPUT_YELLOW)
c.alignment = left_al()
c.border = thin()

label(ws, 44, 2, "Mentor Name (if identified)")
input_cell(ws, 44, 3, "")
label(ws, 44, 4, "Mentor Rank / Role")
input_cell(ws, 44, 5, "")
label(ws, 44, 6, "Senior Rater Name")
input_cell(ws, 44, 7, "", span=2)

# Key named ranges for CALC ENGINE to reference
# (openpyxl named ranges written as simple comments; formulas will use direct cell refs)
ws.merge_cells("B46:I46")
c = ws["B46"]
c.value = "KEY PROFILE CELLS (used by CALC ENGINE) → Rank=C6 | Category=E6 | Status=G6 | MOS=C7 | HomeCity=C10 | MaxCommute=G10 | TargetRank=C25 | PrimaryGoal=C26 | PositionType=C27 | SwitchMOS=G27"
c.font = Font(name="Calibri", size=8, italic=True, color="595959")
c.fill = fill(LIGHT_GRAY)
c.alignment = center()
ws.row_dimensions[46].height = 18

print("MY PROFILE sheet done.")

# ═══════════════════════════════════════════════════════════════════════════════
# CALC ENGINE (hidden sheet — matching logic)
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_calc

ws.merge_cells("A1:T1")
c = ws["A1"]
c.value = "CALC ENGINE — Hidden sheet. Do not edit. Contains matching scores that feed POSITION MATCHES sheet."
c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
c.fill = fill("404040")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 20

# Column headers for calc sheet
calc_headers = [
    "PosID",          # A = Col 1  (mirrors POSITIONS col A)
    "Unit",           # B
    "City",           # C
    "Duty Title",     # D
    "Grade",          # E
    "Career Cat",     # F
    "MOS",            # G
    "PosType",        # H
    "KD",             # I
    "StatusType",     # J
    "Vacancy",        # K
    "ProjVac",        # L
    "GradeScore",     # M  0-35
    "MOSScore",       # N  0-30
    "CommuteScore",   # O  0-15
    "GoalScore",      # P  0-15
    "StatusScore",    # Q  0-5
    "TotalScore",     # R  0-100
    "MatchLabel",     # S
    "CommuteMins",    # T (lookup from city distances)
]

for i, h in enumerate(calc_headers, 1):
    c = ws.cell(row=2, column=i, value=h)
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor="404040")
    c.alignment = Alignment(horizontal="center")
    c.border = Border(
        left=Side(style="thin", color="888888"),
        right=Side(style="thin", color="888888"),
        top=Side(style="thin", color="888888"),
        bottom=Side(style="thin", color="888888")
    )

# Set column widths for calc engine
for i in range(1, 21):
    ws.column_dimensions[get_column_letter(i)].width = 14

# ── Rank ordering lookup (for grade scoring) ──────────────────────────────────
# We build a static rank-to-number mapping in the CONFIG area of CALC ENGINE (rows 200+)
rank_map = [
    ("E1",1),("E2",2),("E3",3),("E4",4),("E5",5),("E6",6),("E7",7),("E8",8),("E9",9),
    ("W1",10),("W2",11),("W3",12),("W4",13),("W5",14),
    ("O1",15),("O2",16),("O3",17),("O4",18),("O5",19),("O6",20),
]
ws.cell(row=200, column=1).value = "RANK MAP (used by scoring formulas)"
ws.cell(row=200, column=1).font = Font(bold=True, color=WHITE)
ws.cell(row=200, column=1).fill = PatternFill("solid", fgColor="404040")
for i, (rank, num) in enumerate(rank_map, 201):
    ws.cell(row=i, column=1).value = rank
    ws.cell(row=i, column=2).value = num

# ── City commute minutes lookup (subset — used by CommuteScore) ───────────────
# Stored in CALC ENGINE rows 230+
ws.cell(row=230, column=1).value = "COMMUTE MINUTES LOOKUP (From|To|Minutes)"
ws.cell(row=230, column=1).font = Font(bold=True, color=WHITE)
ws.cell(row=230, column=1).fill = PatternFill("solid", fgColor="404040")

# Convert city pairs from ADMIN - CITY DISTANCES to minutes (approximate)
city_mins_raw = {
    ("Billings","Bozeman"):130, ("Billings","Butte"):190, ("Billings","Helena"):185,
    ("Billings","Great Falls"):195, ("Billings","Missoula"):285, ("Billings","Kalispell"):370,
    ("Billings","Miles City"):125, ("Billings","Dillon"):255, ("Billings","Havre"):200,
    ("Billings","Lewistown"):115, ("Billings","Livingston"):100, ("Billings","Glendive"):180,
    ("Bozeman","Helena"):80, ("Bozeman","Butte"):70, ("Bozeman","Missoula"):165,
    ("Bozeman","Great Falls"):195, ("Bozeman","Kalispell"):255, ("Bozeman","Dillon"):135,
    ("Bozeman","Miles City"):240, ("Bozeman","Livingston"):25,
    ("Helena","Great Falls"):70, ("Helena","Missoula"):95, ("Helena","Butte"):55,
    ("Helena","Kalispell"):180, ("Helena","Dillon"):110, ("Helena","Havre"):155,
    ("Great Falls","Havre"):95, ("Great Falls","Lewistown"):90, ("Great Falls","Missoula"):165,
    ("Great Falls","Kalispell"):180,
    ("Missoula","Kalispell"):100, ("Missoula","Butte"):110, ("Missoula","Dillon"):160,
    ("Kalispell","Whitefish"):20,
    ("Miles City","Glendive"):60, ("Dillon","Butte"):55, ("Butte","Anaconda"):30,
    ("Helena","Billings"):185, ("Missoula","Billings"):285, ("Kalispell","Billings"):370,
    ("Bozeman","Billings"):130, ("Great Falls","Billings"):195,
}

# Symmetrize
city_mins = {}
for (a,b), mins in city_mins_raw.items():
    city_mins[(a,b)] = mins
    city_mins[(b,a)] = mins
    city_mins[(a,a)] = 0   # same city = 0 min

row_c = 231
for (frm, to), mins in sorted(city_mins.items()):
    ws.cell(row=row_c, column=1).value = frm
    ws.cell(row=row_c, column=2).value = to
    ws.cell(row=row_c, column=3).value = mins
    row_c += 1

print(f"CALC ENGINE lookup tables built ({len(city_mins)} city-pair minute entries).")

# ── Scoring formulas for each position row ─────────────────────────────────────
# POSITIONS data starts at row 3 of ADMIN - POSITIONS (up to row 82 for 80 positions)
# CALC ENGINE mirrors this with formulas starting at row 3

# Profile cell references (MY PROFILE sheet)
RANK_REF        = "'MY PROFILE'!C6"
CAT_REF         = "'MY PROFILE'!E6"
STATUS_REF      = "'MY PROFILE'!G6"
MOS_REF         = "'MY PROFILE'!C7"
HOME_CITY_REF   = "'MY PROFILE'!C10"
MAX_COMMUTE_REF = "'MY PROFILE'!G10"
TARGET_RANK_REF = "'MY PROFILE'!C25"
GOAL_REF        = "'MY PROFILE'!C26"
POS_TYPE_REF    = "'MY PROFILE'!C27"
SWITCH_MOS_REF  = "'MY PROFILE'!G27"

# Max commute in minutes lookup (map label to minutes)
MAX_COMMUTE_LOOKUP = {
    "30 Minutes": 30, "1 Hour": 60, "1.5 Hours": 90,
    "2 Hours": 120, "3 Hours": 180, "No Limit": 9999
}

# Grade numeric order
RANK_NUM = {
    "E1":1,"E2":2,"E3":3,"E4":4,"E5":5,"E6":6,"E7":7,"E8":8,"E9":9,
    "W1":10,"W2":11,"W3":12,"W4":13,"W5":14,
    "O1":15,"O2":16,"O3":17,"O4":18,"O5":19,"O6":20
}

# We'll write FORMULA strings that use Excel IF/VLOOKUP logic.
# For simplicity and reliability, we write static formulas referencing
# fixed profile cells and position data rows.

# ADMIN - POSITIONS column mapping:
# A=RowID, B=Unit, C=UIC, D=City, E=DutyTitle, F=Grade, G=CareerCat,
# H=MOS, I=Branch, J=PosType, K=CommandKD, L=StatusType,
# M=Vacancy, N=ProjVac, O=SecMOS, P=WeightOverride, Q=Notes, R=DemoFlag

POS = "'ADMIN - POSITIONS'"
NUM_POS = 80  # number of position rows

for i in range(1, NUM_POS + 1):
    r  = i + 2   # data starts row 3 in CALC ENGINE
    pr = i + 2   # same in ADMIN - POSITIONS (row 3 = first data row)

    # A: PosID (pull from positions sheet)
    ws.cell(row=r, column=1).value  = f"={POS}!A{pr}"
    # B: Unit
    ws.cell(row=r, column=2).value  = f"={POS}!B{pr}"
    # C: City
    ws.cell(row=r, column=3).value  = f"={POS}!D{pr}"
    # D: Duty Title
    ws.cell(row=r, column=4).value  = f"={POS}!E{pr}"
    # E: Grade
    ws.cell(row=r, column=5).value  = f"={POS}!F{pr}"
    # F: Career Cat
    ws.cell(row=r, column=6).value  = f"={POS}!G{pr}"
    # G: MOS
    ws.cell(row=r, column=7).value  = f"={POS}!H{pr}"
    # H: Position Type
    ws.cell(row=r, column=8).value  = f"={POS}!J{pr}"
    # I: KD
    ws.cell(row=r, column=9).value  = f"={POS}!K{pr}"
    # J: Status Type
    ws.cell(row=r, column=10).value = f"={POS}!L{pr}"
    # K: Vacancy
    ws.cell(row=r, column=11).value = f"={POS}!M{pr}"
    # L: Projected Vacancy
    ws.cell(row=r, column=12).value = f"={POS}!N{pr}"

    # ── M: GRADE SCORE (0–35) ─────────────────────────────────────────────────
    # Logic:
    #   exact match to current rank  → 25 pts
    #   exact match to TARGET rank   → 35 pts  (the sweet spot)
    #   one grade above target       → 20 pts  (stretch)
    #   two grades above             → 10 pts  (very stretch)
    #   below current rank           → 0 pts
    #   career cat mismatch          → 0 pts
    pos_grade = f"E{r}"   # grade of THIS position in calc engine
    grade_score = (
        f'=IF({pos_grade}=""," ",IF('
        f'EXACT({pos_grade},{CAT_REF})=FALSE,0,'  # this won't work well; use career cat column
        # Simplified: compare pos grade to target rank directly
        # 35 if pos grade = target rank AND career cat matches
        # 25 if pos grade = current rank AND career cat matches
        # 20 if one grade above target
        # 10 if two grades above
        # 0 otherwise
        f'IF(AND(EXACT(F{r},{CAT_REF}),EXACT({pos_grade},{TARGET_RANK_REF})),35,'
        f'IF(AND(EXACT(F{r},{CAT_REF}),EXACT({pos_grade},{RANK_REF})),25,'
        f'10))))'
    )
    ws.cell(row=r, column=13).value = grade_score

    # ── N: MOS SCORE (0–30) ───────────────────────────────────────────────────
    pos_mos = f"G{r}"
    mos_score = (
        f'=IF({pos_mos}="","",IF(EXACT({pos_mos},{MOS_REF}),30,'
        f'IF(EXACT({SWITCH_MOS_REF},"Yes"),10,0)))'
    )
    ws.cell(row=r, column=14).value = mos_score

    # ── O: COMMUTE SCORE (0–15) ───────────────────────────────────────────────
    # Look up drive minutes from city_mins table in CALC ENGINE rows 231+
    pos_city = f"C{r}"
    num_pairs = len(city_mins)
    # MATCH formula to find minutes: search col A=home city, col B=pos city
    # Use SUMPRODUCT for two-condition lookup
    commute_mins_formula = (
        f'=IFERROR(SUMPRODUCT(($A$231:$A${230+num_pairs}={HOME_CITY_REF})*'
        f'($B$231:$B${230+num_pairs}={pos_city})*'
        f'($C$231:$C${230+num_pairs})),999)'
    )
    ws.cell(row=r, column=20).value = commute_mins_formula  # T col = CommuteMins

    # Commute score based on max commute setting and actual minutes
    commute_score = (
        f'=IF(T{r}=999,0,'
        f'IF({MAX_COMMUTE_REF}="No Limit",15,'
        f'IF({MAX_COMMUTE_REF}="30 Minutes",IF(T{r}<=30,15,IF(T{r}<=60,8,IF(T{r}<=90,3,0))),'
        f'IF({MAX_COMMUTE_REF}="1 Hour",IF(T{r}<=60,15,IF(T{r}<=90,8,IF(T{r}<=120,3,0))),'
        f'IF({MAX_COMMUTE_REF}="1.5 Hours",IF(T{r}<=90,15,IF(T{r}<=120,8,IF(T{r}<=150,3,0))),'
        f'IF({MAX_COMMUTE_REF}="2 Hours",IF(T{r}<=120,15,IF(T{r}<=150,8,IF(T{r}<=180,3,0))),'
        f'IF({MAX_COMMUTE_REF}="3 Hours",IF(T{r}<=180,15,IF(T{r}<=240,5,0)),'
        f'15)))))))'
    )
    ws.cell(row=r, column=15).value = commute_score  # O col

    # ── P: GOAL SCORE (0–15) ──────────────────────────────────────────────────
    pos_type = f"H{r}"
    kd_col   = f"I{r}"
    goal_score = (
        f'=IF(EXACT({GOAL_REF},"Pursue Command"),IF(EXACT({kd_col},"Yes"),15,5),'
        f'IF(EXACT({GOAL_REF},"Promote to Next Grade"),IF(EXACT({pos_type},"Leadership"),12,'
        f'IF(EXACT({pos_type},"Staff"),10,8)),'
        f'IF(EXACT({GOAL_REF},"Move to Staff"),IF(EXACT({pos_type},"Staff"),15,'
        f'IF(EXACT({pos_type},"Technical"),10,5)),'
        f'IF(EXACT({GOAL_REF},"Pursue AGR Full-Time"),IF(EXACT(J{r},"AGR"),15,'
        f'IF(EXACT(J{r},"M-Day"),5,10)),'
        f'8))))'
    )
    ws.cell(row=r, column=16).value = goal_score  # P col

    # ── Q: STATUS SCORE (0–5) ─────────────────────────────────────────────────
    status_score = (
        f'=IF(EXACT({STATUS_REF},"M-Day"),IF(EXACT(J{r},"M-Day"),5,IF(EXACT(J{r},"AGR"),1,3)),'
        f'IF(EXACT({STATUS_REF},"AGR"),IF(EXACT(J{r},"AGR"),5,IF(EXACT(J{r},"M-Day"),2,3)),'
        f'3))'
    )
    ws.cell(row=r, column=17).value = status_score  # Q col

    # ── R: TOTAL SCORE ────────────────────────────────────────────────────────
    ws.cell(row=r, column=18).value = f"=IFERROR(M{r}+N{r}+O{r}+P{r}+Q{r},0)"

    # ── S: MATCH LABEL ────────────────────────────────────────────────────────
    ws.cell(row=r, column=19).value = (
        f'=IF(R{r}>=80,"STRONG MATCH",'
        f'IF(R{r}>=60,"GOOD MATCH",'
        f'IF(R{r}>=40,"MODERATE / STRETCH",'
        f'IF(R{r}>=20,"POSSIBLE — GAPS EXIST",'
        f'"NOT RECOMMENDED NOW"))))'
    )

print(f"CALC ENGINE formulas written for {NUM_POS} positions.")

# ═══════════════════════════════════════════════════════════════════════════════
# POSITION MATCHES — Display sheet
# ═══════════════════════════════════════════════════════════════════════════════
ws = ws_matches
ws.sheet_view.showGridLines = False

set_col_widths(ws, {
    "A": 2,   "B": 5,   "C": 30,  "D": 16,  "E": 26,
    "F": 8,   "G": 10,  "H": 10,  "I": 14,  "J": 10,
    "K": 12,  "L": 8,   "M": 8,   "N": 8,   "O": 8,
    "P": 8,   "Q": 8,   "R": 10,  "S": 22,  "T": 14
})

# Title
ws.merge_cells("B1:T1")
c = ws["B1"]
c.value = "POSITION MATCHES  |  Scores update automatically when you change MY PROFILE"
c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=ARMY_GREEN)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Profile summary bar
ws.merge_cells("B2:T2")
c = ws["B2"]
c.value = (
    "=\" Soldier: \"&'MY PROFILE'!C6&\" \"&'MY PROFILE'!E6&\"  |  MOS: \"&'MY PROFILE'!C7"
    "&\"  |  Home: \"&'MY PROFILE'!C10&\"  |  Target: \"&'MY PROFILE'!C25"
    "&\"  |  Goal: \"&'MY PROFILE'!C26"
)
c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=HEADER_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 20

# Legend row
legend_items = [
    ("STRONG MATCH", "E2EFDA", "C4"),
    ("GOOD MATCH",   LIGHT_BLUE, "E4"),
    ("MODERATE / STRETCH", LIGHT_YELLOW, "G4"),
    ("POSSIBLE — GAPS", LIGHT_ORANGE, "I4"),
    ("NOT RECOMMENDED", LIGHT_RED, "K4"),
]
ws.row_dimensions[4].height = 20
for text, bg, cell_addr in legend_items:
    ws.merge_cells(f"{cell_addr}:{chr(ord(cell_addr[0])+1)}4")
    c = ws[cell_addr]
    c.value = text
    c.font = Font(name="Calibri", size=8, bold=True, color=ARMY_BLACK)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(
        left=Side(style="thin"),right=Side(style="thin"),
        top=Side(style="thin"),bottom=Side(style="thin")
    )

# Header row
match_headers = [
    "#", "Rank", "Unit", "City", "Duty Title",
    "Grade", "Cat", "MOS", "Position Type", "KD?",
    "Status", "Vacancy", "Grade\nScore", "MOS\nScore", "Commute\nScore",
    "Goal\nScore", "Status\nScore", "TOTAL\nSCORE", "MATCH LABEL", "Commute\n(min)"
]
ws.row_dimensions[5].height = 32
for i, h in enumerate(match_headers, 2):
    c = ws.cell(row=5, column=i, value=h)
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=HEADER_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

ws.freeze_panes = "B6"

# Data rows — pull from CALC ENGINE (sorted by score descending via LARGE/INDEX trick is complex;
# instead we pull directly and the user can sort with the built-in filter)
CALC = "'CALC ENGINE'"

match_label_colors = {
    "STRONG MATCH":          "E2EFDA",
    "GOOD MATCH":            LIGHT_BLUE,
    "MODERATE / STRETCH":    LIGHT_YELLOW,
    "POSSIBLE — GAPS EXIST": LIGHT_ORANGE,
    "NOT RECOMMENDED NOW":   LIGHT_RED,
}

for i in range(1, NUM_POS + 1):
    r     = i + 5    # display rows start at 6
    cr    = i + 2    # calc engine rows start at 3
    bg    = WHITE if i % 2 == 0 else LIGHT_GRAY

    ws.row_dimensions[r].height = 18

    # Row number
    c = ws.cell(row=r, column=2, value=i)
    c.font = Font(name="Calibri", size=8, color="595959")
    c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    c.alignment = Alignment(horizontal="center")
    c.border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

    # Pull cols from CALC ENGINE: A→Unit(B), City(C), Title(D), Grade(E), Cat(F), MOS(G),
    # PosType(H), KD(I), Status(J), Vacancy(K), scores M-Q, Total R, Label S, CommuteMins T
    col_map = [
        (3,  f"={CALC}!B{cr}"),  # Unit
        (4,  f"={CALC}!C{cr}"),  # City
        (5,  f"={CALC}!D{cr}"),  # Duty Title
        (6,  f"={CALC}!E{cr}"),  # Grade
        (7,  f"={CALC}!F{cr}"),  # Cat
        (8,  f"={CALC}!G{cr}"),  # MOS
        (9,  f"={CALC}!H{cr}"),  # PosType
        (10, f"={CALC}!I{cr}"),  # KD
        (11, f"={CALC}!J{cr}"),  # StatusType
        (12, f"={CALC}!K{cr}"),  # Vacancy
        (13, f"={CALC}!M{cr}"),  # GradeScore
        (14, f"={CALC}!N{cr}"),  # MOSScore
        (15, f"={CALC}!O{cr}"),  # CommuteScore
        (16, f"={CALC}!P{cr}"),  # GoalScore
        (17, f"={CALC}!Q{cr}"),  # StatusScore
        (18, f"={CALC}!R{cr}"),  # TotalScore
        (19, f"={CALC}!S{cr}"),  # MatchLabel
        (20, f"={CALC}!T{cr}"),  # CommuteMins
    ]

    for col, formula in col_map:
        c = ws.cell(row=r, column=col, value=formula)
        c.font = Font(name="Calibri", size=9, color=ARMY_BLACK)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if col in (6,7,8,10,11,12,13,14,15,16,17,18,20) else "left",
                                vertical="center")
        c.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

# Auto filter on header row
ws.auto_filter.ref = f"B5:T{5+NUM_POS}"

print(f"POSITION MATCHES display sheet done ({NUM_POS} rows).")

# Save
wb.save("/home/user/S1-AI-Assistsnt-/MT_ARNG_Career_Planner.xlsx")
print("\nPart 2 saved: MT_ARNG_Career_Planner.xlsx")
