"""
Part 3: COMMUTE ANALYSIS + CAREER TIMELINE + COUNSELING SHEET.
Run AFTER parts 1 and 2.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Shared styles ─────────────────────────────────────────────────────────────
ARMY_GREEN  = "1B4F2A"
ARMY_TAN    = "C8A96E"
ARMY_BLACK  = "1A1A1A"
HEADER_BLUE = "1F3864"
LIGHT_BLUE  = "D9E1F2"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW= "FFFF99"
LIGHT_ORANGE= "FCE4D6"
LIGHT_RED   = "FFE0E0"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
INPUT_YELLOW= "FFFACD"
DEMO_ORANGE = "FF6600"

def thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def med():
    s = Side(style="medium", color=ARMY_GREEN)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_al():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def hdr(ws, row, col, text, bg=HEADER_BLUE, fg=WHITE, size=10, span=1, height=24):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=size, bold=True, color=fg)
    c.fill = fill(bg)
    c.alignment = center()
    c.border = thin()
    ws.row_dimensions[row].height = height
    return c

def cell(ws, row, col, value="", bg=WHITE, fg=ARMY_BLACK, bold=False,
         size=9, align="left", border=True, span=1, height=None):
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=size, bold=bold, color=fg)
    c.fill = fill(bg)
    c.alignment = center() if align == "center" else left_al()
    if border:
        c.border = thin()
    if height:
        ws.row_dimensions[row].height = height
    return c

# Profile refs
RANK   = "'MY PROFILE'!C6"
CAT    = "'MY PROFILE'!E6"
MOS    = "'MY PROFILE'!C7"
HOME   = "'MY PROFILE'!C10"
MCMUTE = "'MY PROFILE'!G10"
TRANK  = "'MY PROFILE'!C25"
GOAL   = "'MY PROFILE'!C26"
CALC   = "'CALC ENGINE'"
NUM_POS = 80

wb = load_workbook("/home/user/S1-AI-Assistsnt-/MT_ARNG_Career_Planner.xlsx")

# ═══════════════════════════════════════════════════════════════════════════════
# COMMUTE ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb["COMMUTE ANALYSIS"]
ws.sheet_view.showGridLines = False

set_col_widths(ws, {
    "A":2, "B":5, "C":30, "D":16, "E":26, "F":8,
    "G":14, "H":14, "I":16, "J":14, "K":14, "L":12
})

hdr(ws, 1, 2, "COMMUTE ANALYSIS  |  Drive time and distance impact for each matched position",
    bg=ARMY_GREEN, size=14, span=11, height=30)

# Dynamic summary bar
ws.merge_cells("B2:L2")
c = ws["B2"]
c.value = f'="Home City: "&{HOME}&"   |  Max Commute: "&{MCMUTE}&"   |  Positions showing: All 80 — filter by MATCH LABEL or VACANCY"'
c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
c.fill = fill(HEADER_BLUE)
c.alignment = left_al()
ws.row_dimensions[2].height = 20

# Legend
legend_data = [
    ("SHORT  (≤30 min)", LIGHT_GREEN,  "C4"),
    ("MEDIUM (31–90 min)", LIGHT_BLUE,  "E4"),
    ("LONG   (91–180 min)", LIGHT_YELLOW, "G4"),
    ("VERY LONG (>180 min)", LIGHT_RED,   "I4"),
    ("SAME CITY", "E2EFDA",              "K4"),
]
ws.row_dimensions[4].height = 18
for text, bg, addr in legend_data:
    ec = chr(ord(addr[0])+1) + addr[1:]
    ws.merge_cells(f"{addr}:{ec}")
    c = ws[addr]
    c.value = text
    c.font = Font(name="Calibri", size=8, bold=True)
    c.fill = fill(bg)
    c.alignment = center()
    c.border = thin()

commute_headers = [
    "#", "Unit", "Position City", "Duty Title", "Grade",
    "One-Way\n(min)", "One-Way\n(miles approx)", "Round Trip\n(min)",
    "Drill WE\n(4 legs, hrs)", "Within\nMax?", "Commute\nCategory", "Match\nLabel"
]
ws.row_dimensions[5].height = 36
for i, h in enumerate(commute_headers, 2):
    c = ws.cell(row=5, column=i, value=h)
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill = fill(HEADER_BLUE)
    c.alignment = center()
    c.border = thin()

ws.freeze_panes = "B6"

# Miles lookup table — store in commute sheet rows 200+
# Use ADMIN - CITY DISTANCES col A=From, B=To, C=Miles
MILES_DATA = [
    ("Billings","Bozeman",144), ("Billings","Butte",223), ("Billings","Helena",221),
    ("Billings","Great Falls",225), ("Billings","Missoula",344), ("Billings","Kalispell",449),
    ("Billings","Miles City",141), ("Billings","Dillon",302), ("Billings","Havre",231),
    ("Billings","Lewistown",126), ("Billings","Livingston",115), ("Billings","Glendive",213),
    ("Bozeman","Helena",95), ("Bozeman","Butte",82), ("Bozeman","Missoula",199),
    ("Bozeman","Great Falls",225), ("Bozeman","Kalispell",303), ("Bozeman","Dillon",157),
    ("Bozeman","Miles City",285), ("Bozeman","Livingston",26),
    ("Helena","Great Falls",91), ("Helena","Missoula",113), ("Helena","Butte",65),
    ("Helena","Kalispell",217), ("Helena","Dillon",124), ("Helena","Havre",183),
    ("Great Falls","Havre",115), ("Great Falls","Lewistown",105), ("Great Falls","Missoula",198),
    ("Great Falls","Kalispell",222), ("Missoula","Kalispell",118), ("Missoula","Butte",129),
    ("Missoula","Dillon",189), ("Kalispell","Whitefish",16), ("Miles City","Glendive",72),
    ("Dillon","Butte",63), ("Butte","Anaconda",25), ("Havre","Lewistown",127),
    ("Havre","Glasgow",171), ("Glasgow","Wolf Point",49), ("Great Falls","Glasgow",228),
]
# Symmetrize
miles_sym = {}
for (a,b,m) in MILES_DATA:
    miles_sym[(a,b)] = m
    miles_sym[(b,a)] = m
    miles_sym[(a,a)] = 0

ws.cell(row=200, column=1).value = "MILES LOOKUP (A=From, B=To, C=Miles)"
for i, ((fr, to), mi) in enumerate(sorted(miles_sym.items()), 201):
    ws.cell(row=i, column=1).value = fr
    ws.cell(row=i, column=2).value = to
    ws.cell(row=i, column=3).value = mi

num_miles = len(miles_sym)

for i in range(1, NUM_POS + 1):
    r  = i + 5
    cr = i + 2   # CALC ENGINE row

    ws.row_dimensions[r].height = 17

    # Row #
    cell(ws, r, 2, i, bg=LIGHT_GRAY, align="center", size=8)

    # Unit, City, Title, Grade from CALC ENGINE
    for col, calc_col in [(3,"B"),(4,"C"),(5,"D"),(6,"E")]:
        c = ws.cell(row=r, column=col, value=f"={CALC}!{calc_col}{cr}")
        c.font = Font(name="Calibri", size=9)
        c.fill = fill(WHITE if i%2==0 else LIGHT_GRAY)
        c.alignment = left_al()
        c.border = thin()

    # Commute minutes (from CALC ENGINE col T)
    commute_min_formula = f"={CALC}!T{cr}"
    c_min = ws.cell(row=r, column=7, value=commute_min_formula)
    c_min.font = Font(name="Calibri", size=9, bold=True)
    c_min.fill = fill(WHITE if i%2==0 else LIGHT_GRAY)
    c_min.alignment = center()
    c_min.border = thin()

    # Miles (lookup from local table rows 201+)
    pos_city = f"'COMMUTE ANALYSIS'!D{r}"  # city is col D here
    miles_formula = (
        f'=IFERROR(SUMPRODUCT(($A$201:$A${200+num_miles}={HOME})*'
        f'($B$201:$B${200+num_miles}=C{r})*'
        f'($C$201:$C${200+num_miles})),0)'
    )
    c_mi = ws.cell(row=r, column=8, value=miles_formula)
    c_mi.font = Font(name="Calibri", size=9)
    c_mi.fill = fill(WHITE if i%2==0 else LIGHT_GRAY)
    c_mi.alignment = center()
    c_mi.border = thin()

    # Round trip minutes
    c_rt = ws.cell(row=r, column=9, value=f"=IF(G{r}=999,\"\",G{r}*2)")
    c_rt.font = Font(name="Calibri", size=9)
    c_rt.fill = fill(WHITE if i%2==0 else LIGHT_GRAY)
    c_rt.alignment = center()
    c_rt.border = thin()

    # Drill weekend burden (4 legs drive + assume 2 drill days = 4 one-way trips)
    c_dw = ws.cell(row=r, column=10,
                   value=f'=IF(G{r}=999,"",ROUND(G{r}*4/60,1)&" hrs")')
    c_dw.font = Font(name="Calibri", size=9)
    c_dw.fill = fill(WHITE if i%2==0 else LIGHT_GRAY)
    c_dw.alignment = center()
    c_dw.border = thin()

    # Within max?
    within = (
        f'=IF(G{r}=999,"Unknown",'
        f'IF({MCMUTE}="No Limit","Yes",'
        f'IF({MCMUTE}="30 Minutes",IF(G{r}<=30,"Yes","No"),'
        f'IF({MCMUTE}="1 Hour",IF(G{r}<=60,"Yes","No"),'
        f'IF({MCMUTE}="1.5 Hours",IF(G{r}<=90,"Yes","No"),'
        f'IF({MCMUTE}="2 Hours",IF(G{r}<=120,"Yes","No"),'
        f'IF({MCMUTE}="3 Hours",IF(G{r}<=180,"Yes","No"),"No")))))))'
    )
    c_w = ws.cell(row=r, column=11, value=within)
    c_w.font = Font(name="Calibri", size=9, bold=True)
    c_w.fill = fill(WHITE if i%2==0 else LIGHT_GRAY)
    c_w.alignment = center()
    c_w.border = thin()

    # Category
    cat_formula = (
        f'=IF(G{r}=0,"Same City",'
        f'IF(G{r}<=30,"Short",'
        f'IF(G{r}<=90,"Medium",'
        f'IF(G{r}<=180,"Long","Very Long"))))'
    )
    c_cat = ws.cell(row=r, column=12, value=cat_formula)
    c_cat.font = Font(name="Calibri", size=9, bold=True)
    c_cat.fill = fill(WHITE if i%2==0 else LIGHT_GRAY)
    c_cat.alignment = center()
    c_cat.border = thin()

    # Match label from CALC
    c_ml = ws.cell(row=r, column=13, value=f"={CALC}!S{cr}")
    c_ml.font = Font(name="Calibri", size=9, bold=True)
    c_ml.fill = fill(WHITE if i%2==0 else LIGHT_GRAY)
    c_ml.alignment = center()
    c_ml.border = thin()

ws.auto_filter.ref = f"B5:M{5+NUM_POS}"
print("COMMUTE ANALYSIS done.")

# ═══════════════════════════════════════════════════════════════════════════════
# CAREER TIMELINE
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb["CAREER TIMELINE"]
ws.sheet_view.showGridLines = False

set_col_widths(ws, {
    "A":2, "B":14, "C":8, "D":22, "E":24, "F":22, "G":22, "H":22, "I":14
})

hdr(ws, 1, 2, "CAREER TIMELINE  |  Suggested 10–20 Year Pathway Based on Your Profile",
    bg=ARMY_GREEN, size=14, span=8, height=30)

ws.merge_cells("B2:I2")
c = ws["B2"]
c.value = (
    f'="Rank: "&{RANK}&"  |  MOS: "&{MOS}&"  |  Goal: "&{TRANK}'
    f'&"  |  Primary Goal: "&{GOAL}'
)
c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
c.fill = fill(HEADER_BLUE)
c.alignment = left_al()
ws.row_dimensions[2].height = 20

ws.merge_cells("B3:I3")
c = ws["B3"]
c.value = (
    "This timeline is a PLANNING GUIDE, not a guarantee. Adjust based on force structure, "
    "available positions, and board outcomes. Positions typically held 2–3 years each."
)
c.font = Font(name="Calibri", size=9, italic=True, color="595959")
c.fill = fill(LIGHT_BLUE)
c.alignment = center()
ws.row_dimensions[3].height = 18

# ── Static enlisted pathway grids ────────────────────────────────────────────
# We build one representative pathway for each career category
# Color codes for phase types
PHASE_COLORS = {
    "School/PME":   "7030A0",  # purple
    "Leadership":   ARMY_GREEN,
    "Staff":        HEADER_BLUE,
    "Command":      "C00000",  # red
    "Broadening":   "BF8F00",  # gold
    "Transition":   "595959",
}

def phase_block(ws, row, col, phase_type, title, subtitle, years, span=1, height=55):
    color = PHASE_COLORS.get(phase_type, HEADER_BLUE)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row+2, end_column=col+span-1)
    else:
        ws.merge_cells(start_row=row, start_column=col, end_row=row+2, end_column=col)
    c = ws.cell(row=row, column=col)
    c.value = f"{title}\n({years} yrs)\n{subtitle}"
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill = fill(color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = med()
    for r2 in range(row, row+3):
        ws.row_dimensions[r2].height = 18
    return c

def label_col(ws, row, col, text, bg=LIGHT_GRAY, bold=True, height=54):
    ws.merge_cells(start_row=row, start_column=col, end_row=row+2, end_column=col)
    c = ws.cell(row=row, column=col)
    c.value = text
    c.font = Font(name="Calibri", size=8, bold=bold, color=ARMY_BLACK)
    c.fill = fill(bg)
    c.alignment = center()
    c.border = thin()
    return c

# ── ENLISTED CAREER PATH ──────────────────────────────────────────────────────
row = 5
hdr(ws, row, 2, "ENLISTED CAREER PATH — Example: 92Y → SFC → MSG (Goal)", bg=ARMY_TAN, fg=ARMY_BLACK, span=8, height=22)

phases_enlisted = [
    # (type, title, subtitle, years, col)
    ("Leadership", "E5 / SGT\nTeam Leader / Supply SGT",  "BLC → ALC Prep", "2–3", 3),
    ("School/PME", "ALC\nAdvanced Leader\nCourse",         "Fort Lee or Regional", "3 wks", 4),
    ("Leadership", "E6 / SSG\nSection NCOIC / SL",         "KD: Lead soldiers", "2–3", 5),
    ("School/PME", "SLC\nSenior Leader\nCourse",           "USASMA competitive", "6 wks", 6),
    ("Staff",      "E7 / SFC\nPlatoon SGT / Staff",        "Broaden experience", "2–3", 7),
    ("Leadership", "E8 / MSG\n1SG or Senior Staff",        "Senior leader track", "3+",  8),
    ("Command",    "E8 1SG\nCompany 1SG",                  "Pinnacle for many", "2–3", 9),
]

for ph in phases_enlisted:
    phase_block(ws, row+1, ph[4], ph[0], ph[1], ph[2], ph[3])

# Arrow row
for col in range(3, 10):
    c = ws.cell(row=row+4, column=col)
    c.value = "─────────────────────────────────────────────────────────────────────→"
    c.font = Font(name="Calibri", size=7, color=ARMY_GREEN)
    c.alignment = center()

row += 6

# ── OFFICER PATH ──────────────────────────────────────────────────────────────
hdr(ws, row, 2, "OFFICER CAREER PATH — Example: 2LT → CPT (Company Command) → MAJ → LTC", bg="2E4057", fg=WHITE, span=8, height=22)

phases_officer = [
    ("Staff",     "O1-O2\nPlatoon Leader\n/ XO",         "BOLC → branch qual", "2–3", 3),
    ("School/PME","CCC\nCaptain's Career\nCourse",         "Branch school", "4 mos", 4),
    ("Command",   "O3 CPT\nCompany\nCommand",              "KD: Pinnacle CPT", "2", 5),
    ("Staff",     "O3 CPT\nBattalion Staff\nS3/S4/S1",    "KD staff tour", "2", 6),
    ("School/PME","ILE\nIntermediate\nLevel Ed",           "CGSC / non-res", "1 yr", 7),
    ("Command",   "O5 LTC\nBattalion\nCommand",            "BN CDR — elite", "2", 8),
    ("Staff",     "O5 LTC\nBrigade Staff /\nJ-Staff",      "Senior staff", "3+", 9),
]
for ph in phases_officer:
    phase_block(ws, row+1, ph[4], ph[0], ph[1], ph[2], ph[3])

row += 6

# ── WARRANT PATH ─────────────────────────────────────────────────────────────
hdr(ws, row, 2, "WARRANT OFFICER PATH — Example: WO1 → CW3 → CW5 (Senior Warrant)", bg="5C4033", fg=WHITE, span=8, height=22)

phases_warrant = [
    ("School/PME", "WOFT / WOBC\nWarrant Officer\nBasic Course",  "IERW if aviator", "6–18 mo", 3),
    ("Leadership", "CW2\nTech Expert\nin MOS",                    "WOAC prep", "3–4", 4),
    ("School/PME", "WOAC\nWarrant Officer\nAdvanced Course",      "MOS branch school", "3 wks", 5),
    ("Leadership", "CW3\nSenior Tech\nAdvisor",                   "WOILE prep", "3–4", 6),
    ("School/PME", "WOILE / WOSSE\nWO Intermediate\nLevel Ed",   "Fort Leavenworth", "3 wks", 7),
    ("Leadership", "CW4\nSenior Warrant\nLeader",                 "Mentor CW1/CW2", "4+", 8),
    ("Command",    "CW5\nChief Warrant\nOfficer 5",               "Rare — senior WO", "3+", 9),
]
for ph in phases_warrant:
    phase_block(ws, row+1, ph[4], ph[0], ph[1], ph[2], ph[3])

row += 6

# ── DECISION GATES ────────────────────────────────────────────────────────────
hdr(ws, row, 2, "KEY DECISION GATES — When to take action", bg=ARMY_GREEN, fg=WHITE, span=8, height=22)

gates = [
    ("E4→E5",        "Enroll in BLC. Ensure SSD1 complete. Ask unit for recommendation letter.", LIGHT_GREEN),
    ("E5→E6",        "Complete ALC. Serve in leadership position. Build OER/NCOER support file.", LIGHT_BLUE),
    ("E6→E7 (SFC)",  "Compete for SLC. Seek platoon sergeant or staff NCOIC. Consider AGR track.", LIGHT_YELLOW),
    ("E7→E8 (MSG)",  "SLC complete. Centralized board. 1SG assignment is key. Build senior rater relationship.", LIGHT_ORANGE),
    ("Command Track","Pre-Command Course required. Build relationship with BN CDR and retention NCO.", LIGHT_RED),
    ("WO Interest",  "Talk to state WO recruiter. SIFT test. Class 1A physical (if aviation). Strong GT score.", "EAD1DC"),
    ("AGR Interest", "Contact state AGR office. Position must be vacant. AGR tour = active duty benefits.", LIGHT_GREEN),
    ("MOS Switch",   "Talk to retention NCO. Check line scores and waivers. Plan MOSQ school time off work.", LIGHT_BLUE),
]

for i, (gate, desc, bg) in enumerate(gates, row+1):
    ws.row_dimensions[i].height = 22
    c = ws.cell(row=i, column=2, value=gate)
    c.font = Font(name="Calibri", size=9, bold=True, color=ARMY_BLACK)
    c.fill = fill(ARMY_TAN)
    c.alignment = center()
    c.border = thin()

    ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=9)
    c = ws.cell(row=i, column=3, value=desc)
    c.font = Font(name="Calibri", size=9, color=ARMY_BLACK)
    c.fill = fill(bg)
    c.alignment = left_al()
    c.border = thin()

row += len(gates) + 2

# Legend box
hdr(ws, row, 2, "TIMELINE LEGEND", bg="595959", fg=WHITE, span=2, height=20)
for i, (ptype, color) in enumerate(PHASE_COLORS.items(), row+1):
    ws.row_dimensions[i].height = 18
    c = ws.cell(row=i, column=2)
    c.value = ptype
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill = fill(color)
    c.alignment = center()
    c.border = thin()

print("CAREER TIMELINE done.")

# ═══════════════════════════════════════════════════════════════════════════════
# COUNSELING SHEET — Printable 1-pager
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb["COUNSELING SHEET"]
ws.sheet_view.showGridLines = False

# Page setup for printing
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToPage  = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight= 1
ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3, footer=0.3)
ws.print_area   = "A1:L55"

set_col_widths(ws, {
    "A":2, "B":14, "C":14, "D":14, "E":12, "F":10,
    "G":12, "H":12, "I":12, "J":10, "K":10, "L":2
})

def cs_row(h=16):
    return h   # just for documentation

# ── HEADER ────────────────────────────────────────────────────────────────────
ws.merge_cells("B1:K1")
c = ws["B1"]
c.value = "MONTANA ARMY NATIONAL GUARD — SOLDIER CAREER COUNSELING SHEET"
c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
c.fill = fill(ARMY_GREEN)
c.alignment = center()
ws.row_dimensions[1].height = 28

ws.merge_cells("B2:K2")
c = ws["B2"]
c.value = (
    f'="As of: "&TEXT(TODAY(),"MMMM DD, YYYY")&"   |   Prepared for: "&{RANK}'
    f'&" "&\'MY PROFILE\'!C5&"   |   MOS: "&{MOS}&"   |   Home: "&{HOME}'
)
c.font = Font(name="Calibri", size=9, color=WHITE)
c.fill = fill(HEADER_BLUE)
c.alignment = center()
ws.row_dimensions[2].height = 16

# ── SECTION 1: Current Snapshot ───────────────────────────────────────────────
hdr(ws, 3, 2, "SECTION 1 — CURRENT SNAPSHOT", bg=ARMY_TAN, fg=ARMY_BLACK, span=10, height=18)

snapshot_labels = [
    ("Rank / Grade",      f"={RANK}",          "Career Category",   f"={CAT}"),
    ("MOS / AOC",         f"={MOS}",           "Component",         f"='MY PROFILE'!G6"),
    ("Unit",              "='MY PROFILE'!C8",   "Unit City",         "='MY PROFILE'!G9"),
    ("Current Title",     "='MY PROFILE'!C9",   "Years of Service",  "='MY PROFILE'!C11"),
    ("TIG (yrs)",         "='MY PROFILE'!E11",  "TIS / Svc Entry",   "='MY PROFILE'!E12"),
    ("Clearance",         "='MY PROFILE'!E42",  "APFT Status",       "='MY PROFILE'!C42"),
]

for i, (l1, v1, l2, v2) in enumerate(snapshot_labels, 4):
    ws.row_dimensions[i].height = 17
    c = ws.cell(row=i, column=2, value=l1)
    c.font = Font(name="Calibri", size=8, bold=True)
    c.fill = fill(LIGHT_GRAY)
    c.alignment = left_al()
    c.border = thin()
    ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    c = ws.cell(row=i, column=3, value=v1)
    c.font = Font(name="Calibri", size=9, color="1F3864")
    c.fill = fill(WHITE)
    c.alignment = left_al()
    c.border = thin()

    c = ws.cell(row=i, column=6, value=l2)
    c.font = Font(name="Calibri", size=8, bold=True)
    c.fill = fill(LIGHT_GRAY)
    c.alignment = left_al()
    c.border = thin()
    ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=11)
    c = ws.cell(row=i, column=7, value=v2)
    c.font = Font(name="Calibri", size=9, color="1F3864")
    c.fill = fill(WHITE)
    c.alignment = left_al()
    c.border = thin()

# ── SECTION 2: PME Status ─────────────────────────────────────────────────────
hdr(ws, 10, 2, "SECTION 2 — PME STATUS", bg="2E4057", fg=WHITE, span=10, height=18)

pme_items = [
    ("BLC", "='MY PROFILE'!C16"), ("ALC", "='MY PROFILE'!C17"),
    ("SLC", "='MY PROFILE'!C18"), ("SMC", "='MY PROFILE'!C19"),
    ("BOLC","='MY PROFILE'!C16"), ("CCC", "='MY PROFILE'!C17"),
    ("ILE", "='MY PROFILE'!C18"), ("SSC", "='MY PROFILE'!C19"),
    ("WOBC","='MY PROFILE'!G16"), ("WOAC","='MY PROFILE'!G17"),
]
ws.row_dimensions[11].height = 17
for i, (pme, val) in enumerate(pme_items):
    col_s = 2 + i    # spread across cols B–K
    c = ws.cell(row=11, column=col_s, value=pme)
    c.font = Font(name="Calibri", size=7, bold=True)
    c.fill = fill(LIGHT_GRAY)
    c.alignment = center()
    c.border = thin()

ws.row_dimensions[12].height = 17
for i, (pme, val) in enumerate(pme_items):
    col_s = 2 + i
    c = ws.cell(row=12, column=col_s, value=val)
    c.font = Font(name="Calibri", size=8, bold=True, color="1F3864")
    c.fill = fill(WHITE)
    c.alignment = center()
    c.border = thin()

# ── SECTION 3: Career Goals ───────────────────────────────────────────────────
hdr(ws, 13, 2, "SECTION 3 — CAREER GOALS", bg=ARMY_GREEN, fg=WHITE, span=10, height=18)

goals = [
    ("Target Rank",    f"={TRANK}",           "Primary Goal",      f"={GOAL}"),
    ("Goal Timeline",  "='MY PROFILE'!E25",   "Preferred Status",  "='MY PROFILE'!G25"),
    ("Position Type",  "='MY PROFILE'!C27",   "Open to Command?",  "='MY PROFILE'!E27"),
    ("Switch MOS?",    "='MY PROFILE'!G27",   "WO Interest?",      "='MY PROFILE'!E28"),
    ("Willing to Relocate?","='MY PROFILE'!G12","Max Commute",     f"={MCMUTE}"),
]

for i, (l1, v1, l2, v2) in enumerate(goals, 14):
    ws.row_dimensions[i].height = 17
    c = ws.cell(row=i, column=2, value=l1)
    c.font = Font(name="Calibri", size=8, bold=True)
    c.fill = fill(LIGHT_GREEN)
    c.alignment = left_al()
    c.border = thin()
    ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    c = ws.cell(row=i, column=3, value=v1)
    c.font = Font(name="Calibri", size=9, color="1F3864")
    c.fill = fill(WHITE)
    c.alignment = left_al()
    c.border = thin()

    c = ws.cell(row=i, column=6, value=l2)
    c.font = Font(name="Calibri", size=8, bold=True)
    c.fill = fill(LIGHT_GREEN)
    c.alignment = left_al()
    c.border = thin()
    ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=11)
    c = ws.cell(row=i, column=7, value=v2)
    c.font = Font(name="Calibri", size=9, color="1F3864")
    c.fill = fill(WHITE)
    c.alignment = left_al()
    c.border = thin()

# 5-year goal narrative
r = 19
ws.row_dimensions[r].height = 17
c = ws.cell(row=r, column=2, value="5-Year Goal Statement")
c.font = Font(name="Calibri", size=8, bold=True)
c.fill = fill(LIGHT_GREEN)
c.alignment = left_al()
c.border = thin()
ws.merge_cells(f"C{r}:K{r}")
c = ws.cell(row=r, column=3, value="='MY PROFILE'!B31")
c.font = Font(name="Calibri", size=8, italic=True)
c.fill = fill(WHITE)
c.alignment = left_al()
c.border = thin()

# ── SECTION 4: Top Position Matches ──────────────────────────────────────────
hdr(ws, 21, 2, "SECTION 4 — TOP POSITION MATCHES (from POSITION MATCHES sheet)", bg=HEADER_BLUE, fg=WHITE, span=10, height=18)

# Column headers
match_cols = ["Rank","Unit","City","Duty Title","Grade","Status","Score","Match","Commute"]
ws.row_dimensions[22].height = 20
for i, h in enumerate(match_cols, 2):
    c = ws.cell(row=22, column=i, value=h)
    c.font = Font(name="Calibri", size=8, bold=True, color=WHITE)
    c.fill = fill(HEADER_BLUE)
    c.alignment = center()
    c.border = thin()
# Note: col L is margin so last match col is K
# Merge title+unit across wider space
ws.cell(row=22, column=6).value = "Duty Title"
ws.merge_cells("F22:H22")

# Top 5 matches — static pull from POSITION MATCHES rows 6–10
# (In a real deployment you'd sort POSITION MATCHES by score first)
for i in range(5):
    r = 23 + i
    ws.row_dimensions[r].height = 16
    pm_row = 6 + i   # POSITION MATCHES rows 6–85 (first 5 for counseling sheet)
    PM = "'POSITION MATCHES'"

    vals = [
        (2,  f"={PM}!C{pm_row}"),   # Unit
        (3,  f"={PM}!D{pm_row}"),   # City
        (6,  f"={PM}!E{pm_row}"),   # Duty Title (merged F–H)
        (5,  f"={PM}!F{pm_row}"),   # Grade
        (9,  f"={PM}!K{pm_row}"),   # Status
        (10, f"={PM}!R{pm_row}"),   # Score
        (11, f"={PM}!S{pm_row}"),   # Match label
    ]
    for col, formula in vals:
        c = ws.cell(row=r, column=col, value=formula)
        c.font = Font(name="Calibri", size=8, color=ARMY_BLACK)
        c.fill = fill(WHITE if i%2==0 else LIGHT_GRAY)
        c.alignment = center()
        c.border = thin()

# ── SECTION 5: Gaps to Close ──────────────────────────────────────────────────
hdr(ws, 29, 2, "SECTION 5 — GAPS TO CLOSE (Complete before next board/assignment)", bg="5C4033", fg=WHITE, span=10, height=18)

gap_rows = [
    ("PME Required",        "='MY PROFILE'!C17",  "Status",  "='MY PROFILE'!C17"),
    ("Schools Recommended", "SLC — see PME sheet", "Timeline","Within 2 years"),
    ("MOS Qualification",   "N/A — current MOS qualified", "Action", "Maintain currency"),
    ("Deployment / OCONUS", "='MY PROFILE'!G42",  "Impact",  "Points toward senior boards"),
    ("Education Goal",      "='MY PROFILE'!C22",  "Gap",     "Complete bachelor's degree"),
]
for i, (l1, v1, l2, v2) in enumerate(gap_rows, 30):
    ws.row_dimensions[i].height = 16
    c = ws.cell(row=i, column=2, value=l1)
    c.font = Font(name="Calibri", size=8, bold=True)
    c.fill = fill(LIGHT_ORANGE)
    c.alignment = left_al()
    c.border = thin()
    ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    c = ws.cell(row=i, column=3, value=v1)
    c.font = Font(name="Calibri", size=8)
    c.fill = fill(WHITE if i%2==0 else LIGHT_YELLOW)
    c.alignment = left_al()
    c.border = thin()

    c = ws.cell(row=i, column=6, value=l2)
    c.font = Font(name="Calibri", size=8, bold=True)
    c.fill = fill(LIGHT_ORANGE)
    c.alignment = left_al()
    c.border = thin()
    ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=11)
    c = ws.cell(row=i, column=7, value=v2)
    c.font = Font(name="Calibri", size=8, italic=True)
    c.fill = fill(WHITE if i%2==0 else LIGHT_YELLOW)
    c.alignment = left_al()
    c.border = thin()

# ── SECTION 6: 3-Step Timeline ────────────────────────────────────────────────
hdr(ws, 36, 2, "SECTION 6 — SUGGESTED NEAR-TERM CAREER PATHWAY", bg=ARMY_GREEN, fg=WHITE, span=10, height=18)

steps = [
    ("NOW (0–2 yrs)",    "Complete ALC. Stay in current position. Build NCOER support file. Target E6 board.", LIGHT_GREEN),
    ("MID (2–5 yrs)",    "Move to Squad Leader / NCOIC role. Attend SLC. Consider AGR conversion opportunity.", LIGHT_BLUE),
    ("GOAL (5–10 yrs)",  "Compete for SFC. Seek platoon sergeant billet. Position near Helena or AGR statewide.", LIGHT_ORANGE),
]
for i, (label, text, bg) in enumerate(steps, 37):
    ws.row_dimensions[i].height = 22
    c = ws.cell(row=i, column=2, value=label)
    c.font = Font(name="Calibri", size=8, bold=True, color=WHITE)
    c.fill = fill(ARMY_GREEN if i==37 else HEADER_BLUE if i==38 else ARMY_TAN)
    c.alignment = center()
    c.border = thin()
    ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=11)
    c = ws.cell(row=i, column=3, value=text)
    c.font = Font(name="Calibri", size=8)
    c.fill = fill(bg)
    c.alignment = left_al()
    c.border = thin()

# ── SECTION 7: Discussion Points ──────────────────────────────────────────────
hdr(ws, 40, 2, "SECTION 7 — DISCUSSION POINTS FOR THIS COUNSELING SESSION", bg=HEADER_BLUE, fg=WHITE, span=10, height=18)

discussion = [
    "1.  Next promotion board — Is the Soldier on track? What does the senior rater need to see?",
    "2.  Position of Interest — Are the matched positions realistic? Is command/KD needed?",
    "3.  AGR opportunity — Is the Soldier interested? Any open AGR billets to discuss?",
    "4.  PME gaps — When will the Soldier attend the next required school?",
    "5.  Commute and relocation — Are geographic constraints realistic for stated goals?",
    "6.  MOS switch or branch transfer — Is this being considered? What are the steps?",
    "7.  Warrant Officer program — Any interest? SIFT score and physical status?",
    "8.  Senior rater support — Does the Soldier understand what the senior rater needs to say?",
    "9.  ETS / retention planning — Does the Soldier plan to re-enlist? What incentives exist?",
    "10. Mentorship — Who is the Soldier's mentor? Is there a plan for regular touchpoints?",
]
for i, text in enumerate(discussion, 41):
    ws.row_dimensions[i].height = 15
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=11)
    c = ws.cell(row=i, column=2, value=text)
    c.font = Font(name="Calibri", size=8)
    c.fill = fill(WHITE if i%2==0 else LIGHT_BLUE)
    c.alignment = left_al()
    c.border = thin()

# ── SIGNATURES ────────────────────────────────────────────────────────────────
hdr(ws, 52, 2, "SIGNATURES", bg=ARMY_TAN, fg=ARMY_BLACK, span=10, height=18)

ws.row_dimensions[53].height = 22
sig_items = [
    (2, "Soldier Signature / Date"),
    (5, "Mentor / Rater Signature / Date"),
    (8, "Senior Rater Signature / Date"),
]
for col, text in sig_items:
    ws.merge_cells(start_row=53, start_column=col, end_row=53, end_column=col+2)
    c = ws.cell(row=53, column=col, value=text)
    c.font = Font(name="Calibri", size=8, color="595959", italic=True)
    c.fill = fill(WHITE)
    c.alignment = Alignment(horizontal="left", vertical="bottom")
    c.border = Border(bottom=Side(style="medium", color=ARMY_BLACK))

ws.row_dimensions[54].height = 18
ws.merge_cells("B54:K54")
c = ws["B54"]
c.value = "MTARNG Career Planning Tool v1.0  |  DEMO DATA — Replace admin reference sheets with real MTARNG data before operational use"
c.font = Font(name="Calibri", size=7, italic=True, color="595959")
c.alignment = center()

print("COUNSELING SHEET done.")

wb.save("/home/user/S1-AI-Assistsnt-/MT_ARNG_Career_Planner.xlsx")
print("\nPart 3 saved: MT_ARNG_Career_Planner.xlsx")
print("All sheets complete.")
